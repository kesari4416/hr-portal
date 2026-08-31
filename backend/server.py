from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
import aiomysql
import os
import logging
import json
from pathlib import Path
from pydantic import BaseModel, EmailStr
from typing import List, Optional
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import io
import asyncio

# Cookie security - detect HTTPS from FRONTEND_URL
IS_HTTPS = os.environ.get("FRONTEND_URL", "").startswith("https")
COOKIE_SECURE = IS_HTTPS
COOKIE_SAMESITE = "none" if IS_HTTPS else "lax"
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image as RLImage
from reportlab.lib.units import inch, mm
from num2words import num2words
from openpyxl import Workbook
import httpx
import math
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT_DIR = Path(__file__).parent

# ── Image Storage via MySQL ────────────────────────────────────────────────────
import uuid as _uuid
import base64

MIME_TYPES = {
    "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
    "gif": "image/gif", "webp": "image/webp"
}

async def _save_image_db(data: bytes, filename: str) -> str:
    """Store image bytes in the media table, return unique media_id."""
    media_id = _uuid.uuid4().hex
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "jpg"
    content_type = MIME_TYPES.get(ext, "image/jpeg")
    await execute_query(
        "INSERT INTO media (media_id, filename, content_type, data) VALUES (%s, %s, %s, %s)",
        (media_id, filename, content_type, data)
    )
    return media_id

async def _get_image_db(media_id: str) -> tuple:
    """Retrieve image bytes and content_type from media table."""
    row = await execute_query("SELECT data, content_type FROM media WHERE media_id = %s", (media_id,), fetch_one=True)
    if not row:
        raise FileNotFoundError(f"Media not found: {media_id}")
    return bytes(row["data"]), row["content_type"]
# ─────────────────────────────────────────────────────────────────────────────

# MySQL connection pool
pool = None

async def get_pool():
    global pool
    if pool is None:
        pool = await aiomysql.create_pool(
            host=os.environ.get("MYSQL_HOST", "localhost"),
            port=int(os.environ.get("MYSQL_PORT", 3306)),
            user=os.environ.get("MYSQL_USER", "hruser"),
            password=os.environ.get("MYSQL_PASSWORD", "hrpass123"),
            db=os.environ.get("MYSQL_DB", "hr_portal"),
            autocommit=True,
            minsize=2,
            maxsize=10,
            charset='utf8mb4'
        )
    return pool

async def execute_query(query, args=None, fetch_one=False, fetch_all=False, last_id=False):
    p = await get_pool()
    async with p.acquire() as conn:
        async with conn.cursor(aiomysql.DictCursor) as cur:
            await cur.execute(query, args)
            if last_id:
                return cur.lastrowid
            if fetch_one:
                return await cur.fetchone()
            if fetch_all:
                return await cur.fetchall()
            return cur.rowcount

# JWT Config
JWT_ALGORITHM = "HS256"

def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]

# Password utilities
def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password.encode("utf-8"), salt)
    return hashed.decode("utf-8")

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode("utf-8"), hashed_password.encode("utf-8"))

# JWT utilities
def create_access_token(user_id: int, email: str, role: str) -> str:
    payload = {
        "sub": str(user_id),
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=60),
        "type": "access"
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: int) -> str:
    payload = {
        "sub": str(user_id),
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "refresh"
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

# Auth dependency
async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await execute_query(
            "SELECT id, email, name, role, department, position, avatar_url, casual_leave, sick_leave, loss_of_pay, permission_hours, half_day_leave, shift, basic_salary FROM users WHERE id = %s",
            (int(payload["sub"]),), fetch_one=True
        )
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user["_id"] = str(user["id"])
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def require_admin(request: Request) -> dict:
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

async def require_admin_or_manager(request: Request) -> dict:
    user = await get_current_user(request)
    if user.get("role") not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Admin or Manager access required")
    return user

# Create the main app
app = FastAPI()

# Create routers
api_router = APIRouter(prefix="/api")
auth_router = APIRouter(prefix="/auth")
employees_router = APIRouter(prefix="/employees")
attendance_router = APIRouter(prefix="/attendance")
leave_router = APIRouter(prefix="/leave")
admin_router = APIRouter(prefix="/admin")
permission_router = APIRouter(prefix="/permission")
payslip_router = APIRouter(prefix="/payslip")
reports_router = APIRouter(prefix="/reports")
holidays_router = APIRouter(prefix="/holidays")
policy_router = APIRouter(prefix="/policy")
wfh_router = APIRouter(prefix="/wfh")
cr_router = APIRouter(prefix="/cr")

# Constants
REQUIRED_WORK_HOURS = 8
TOTAL_WORK_HOURS = 8.5
MONTHLY_PERMISSION_HOURS = 2
MAX_PERMISSION_PER_USE = 1
SHORT_DAYS_FOR_HALF_LEAVE = 3
MAX_BREAK_MINUTES = 40
WORKING_DAYS_PER_MONTH = 22

# Holiday List
HOLIDAYS = [
    {"date": "2026-01-01", "day": "Thursday", "festival": "New Year"},
    {"date": "2026-01-26", "day": "Monday", "festival": "Republic Day"},
    {"date": "2026-04-03", "day": "Friday", "festival": "Good Friday"},
    {"date": "2026-04-14", "day": "Tuesday", "festival": "Vishu"},
    {"date": "2026-05-01", "day": "Friday", "festival": "May Day"},
    {"date": "2026-05-27", "day": "Wednesday", "festival": "Bakrid (Tentative)"},
    {"date": "2026-08-15", "day": "Saturday", "festival": "Independence Day"},
    {"date": "2026-08-25", "day": "Tuesday", "festival": "Onam"},
    {"date": "2026-10-02", "day": "Friday", "festival": "Gandhi Jayanti"},
    {"date": "2026-10-20", "day": "Tuesday", "festival": "Vijayadasami"},
    {"date": "2026-11-08", "day": "Sunday", "festival": "Diwali"},
    {"date": "2026-12-25", "day": "Friday", "festival": "Christmas"},
]

def is_weekend(date_str: str) -> bool:
    d = datetime.strptime(date_str, "%Y-%m-%d")
    return d.weekday() in (5, 6)

def is_holiday(date_str: str) -> bool:
    return any(h["date"] == date_str for h in HOLIDAYS)

def get_holiday_name(date_str: str) -> str:
    for h in HOLIDAYS:
        if h["date"] == date_str:
            return h["festival"]
    return ""

SHIFTS = {}  # Legacy - no longer used. Shifts are stored as "HH:MM-HH:MM" in users.shift column

DEFAULT_SHIFT_START = "09:00"
DEFAULT_SHIFT_END = "18:00"

def parse_shift(shift_str: str) -> dict:
    """Parse shift string 'HH:MM-HH:MM' into start/end dict."""
    if shift_str and "-" in shift_str:
        parts = shift_str.split("-", 1)
        if len(parts) == 2:
            return {"start": parts[0].strip(), "end": parts[1].strip()}
    return {"start": DEFAULT_SHIFT_START, "end": DEFAULT_SHIFT_END}

# Geofencing defaults (can be overridden in DB settings table)
DEFAULT_OFFICE_LAT = 10.0159  # Kochi, Kerala
DEFAULT_OFFICE_LNG = 76.3419
DEFAULT_OFFICE_RADIUS_KM = 0.5  # 500 meters

def haversine_km(lat1, lon1, lat2, lon2):
    """Calculate distance between two GPS points in kilometers."""
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    return R * 2 * math.asin(math.sqrt(a))

async def reverse_geocode(lat, lng):
    """Get address from coordinates using Nominatim (free)."""
    try:
        async with httpx.AsyncClient(timeout=5) as client:
            resp = await client.get(
                f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lng}&format=json&addressdetails=1",
                headers={"User-Agent": "SparkCurv-HR-Portal/1.0"}
            )
            if resp.status_code == 200:
                data = resp.json()
                return data.get("display_name", f"{lat}, {lng}")
    except Exception:
        pass
    return f"{lat}, {lng}"

async def get_office_settings():
    """Get office location settings from DB or use defaults."""
    settings = await execute_query("SELECT * FROM office_settings WHERE id = 1", fetch_one=True)
    if settings:
        return {
            "latitude": float(settings["latitude"]),
            "longitude": float(settings["longitude"]),
            "radius_km": float(settings["radius_km"]),
            "name": settings.get("office_name", "Office")
        }
    return {"latitude": DEFAULT_OFFICE_LAT, "longitude": DEFAULT_OFFICE_LNG, "radius_km": DEFAULT_OFFICE_RADIUS_KM, "name": "Office"}

async def check_geofence(lat, lng, user_id):
    """Check if location is within office geofence or user has approved WFH."""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    # Check if user has approved WFH for today
    wfh = await execute_query(
        "SELECT id FROM wfh_requests WHERE user_id = %s AND date = %s AND status = 'approved'",
        (user_id, today), fetch_one=True
    )
    if wfh:
        return True, "WFH"
    # Check office geofence
    office = await get_office_settings()
    distance = haversine_km(lat, lng, office["latitude"], office["longitude"])
    if distance <= office["radius_km"]:
        return True, "Office"
    return False, f"Outside office area ({distance:.1f} km away, max {office['radius_km']} km)"

# Pydantic Models
class UserRegister(BaseModel):
    email: EmailStr
    password: str
    name: str
    department: Optional[str] = "General"
    position: Optional[str] = "Employee"
    role: Optional[str] = "employee"
    employee_code: Optional[str] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class LeaveRequest(BaseModel):
    leave_type: str
    start_date: str
    end_date: str
    reason: str
    is_half_day: bool = False

class EmployeeUpdate(BaseModel):
    name: Optional[str] = None
    department: Optional[str] = None
    position: Optional[str] = None
    casual_leave: Optional[float] = None
    sick_leave: Optional[float] = None
    permission_hours: Optional[float] = None
    shift: Optional[str] = None
    role: Optional[str] = None
    wfh_limit: Optional[int] = None
    employee_code: Optional[str] = None

class PermissionRequest(BaseModel):
    duration_minutes: int
    reason: str
    date: str

class LocationData(BaseModel):
    latitude: float
    longitude: float

class OfficeSettingsUpdate(BaseModel):
    latitude: float
    longitude: float
    radius_km: float = 0.5
    office_name: str = "Office"

class SalaryUpdate(BaseModel):
    basic_salary: float

class PayslipGenerate(BaseModel):
    employee_id: str
    month: int
    year: int

class BulkPayrollGenerate(BaseModel):
    month: int
    year: int

class CustomDeductionCreate(BaseModel):
    user_id: int
    deduction_name: str
    amount: float = 0
    is_percentage: bool = False
    percentage: float = 0
    is_active: bool = True

class ShiftAssign(BaseModel):
    start_time: str  # HH:MM
    end_time: str    # HH:MM

class PasswordReset(BaseModel):
    new_password: str

class WFHRequest(BaseModel):
    date: str
    reason: str

class PolicyCreate(BaseModel):
    title: str
    category: str
    content: str
    icon: Optional[str] = "article"
    sort_order: Optional[int] = 0

class PolicyUpdate(BaseModel):
    title: Optional[str] = None
    category: Optional[str] = None
    content: Optional[str] = None
    icon: Optional[str] = None
    sort_order: Optional[int] = None

class CRCreate(BaseModel):
    title: str
    description: str
    cr_type: str = "General"
    priority: str = "medium"
    metadata: Optional[dict] = None

class CRUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    cr_type: Optional[str] = None
    priority: Optional[str] = None

# ============== AUTH ROUTES ==============

@auth_router.post("/register")
async def register(user_data: UserRegister, response: Response):
    email = user_data.email.lower()
    existing = await execute_query("SELECT id FROM users WHERE email = %s", (email,), fetch_one=True)
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")

    hashed = hash_password(user_data.password)
    user_id = await execute_query(
        """INSERT INTO users (email, password_hash, name, role, department, position, avatar_url, created_at, casual_leave, sick_leave, loss_of_pay, permission_hours, half_day_leave, shift)
           VALUES (%s, %s, %s, %s, %s, %s, '', %s, 12, 3, 0, %s, 0, '')""",
        (email, hashed, user_data.name, "employee", user_data.department, user_data.position, datetime.now(timezone.utc).isoformat(), MONTHLY_PERMISSION_HOURS),
        last_id=True
    )

    access_token = create_access_token(user_id, email, "employee")
    refresh_token = create_refresh_token(user_id)
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=COOKIE_SECURE, samesite=COOKIE_SAMESITE, max_age=3600, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=COOKIE_SECURE, samesite=COOKIE_SAMESITE, max_age=604800, path="/")

    return {"id": str(user_id), "email": email, "name": user_data.name, "role": "employee", "department": user_data.department, "position": user_data.position, "avatar_url": ""}

@auth_router.post("/login")
async def login(user_data: UserLogin, response: Response):
    email = user_data.email.lower()
    user = await execute_query("SELECT * FROM users WHERE email = %s", (email,), fetch_one=True)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    if not verify_password(user_data.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")

    access_token = create_access_token(user["id"], email, user["role"])
    refresh_token = create_refresh_token(user["id"])
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=COOKIE_SECURE, samesite=COOKIE_SAMESITE, max_age=3600, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=COOKIE_SECURE, samesite=COOKIE_SAMESITE, max_age=604800, path="/")

    return {"id": str(user["id"]), "email": email, "name": user["name"], "role": user["role"], "department": user.get("department", "General"), "position": user.get("position", "Employee"), "avatar_url": user.get("avatar_url"), "employee_code": user.get("employee_code", "")}

@auth_router.post("/logout")
async def logout(response: Response):
    response.delete_cookie(key="access_token", path="/")
    response.delete_cookie(key="refresh_token", path="/")
    return {"message": "Logged out successfully"}

@auth_router.get("/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    user.pop("password_hash", None)
    user.pop("basic_salary", None)
    return user

@auth_router.get("/password-reset-info")
async def password_reset_info():
    return {"message": "To reset your password, please contact HR department.", "contact_email": "hr@company.com", "contact_method": "Email HR or visit HR office during working hours"}

# ============== ATTENDANCE ROUTES ==============

@attendance_router.post("/clock-in")
async def clock_in(request: Request):
    user = await get_current_user(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    if is_weekend(today):
        raise HTTPException(status_code=400, detail="Cannot clock in on weekends (Saturday/Sunday)")
    if is_holiday(today):
        raise HTTPException(status_code=400, detail=f"Cannot clock in on holiday: {get_holiday_name(today)}")

    # Parse location from request body
    try:
        body = await request.json()
        lat = body.get("latitude")
        lng = body.get("longitude")
    except Exception:
        lat = None
        lng = None

    # Check if user has approved WFH today — GPS becomes optional
    wfh_today = await execute_query(
        "SELECT id FROM wfh_requests WHERE user_id = %s AND date = %s AND status = 'approved'",
        (user["id"], today), fetch_one=True
    )
    is_wfh = bool(wfh_today)

    if lat is None or lng is None:
        if not is_wfh:
            raise HTTPException(status_code=400, detail="Location is required. Please enable GPS/location services.")
        # WFH employee — clock in without GPS
        lat = 0.0
        lng = 0.0

    # Geofence check (will auto-pass for WFH employees)
    allowed, location_type = await check_geofence(lat, lng, user["id"])
    if not allowed:
        raise HTTPException(status_code=400, detail=f"Cannot clock in: {location_type}. You must be at the office or have approved WFH.")

    existing = await execute_query(
        "SELECT id FROM attendance WHERE user_id = %s AND date = %s AND clock_out IS NULL", (user["id"], today), fetch_one=True
    )
    if existing:
        raise HTTPException(status_code=400, detail="Already clocked in")

    # Reverse geocode (skip for WFH with no location)
    address = await reverse_geocode(lat, lng) if (lat != 0.0 or lng != 0.0) else "Work From Home"

    att_id = await execute_query(
        "INSERT INTO attendance (user_id, user_name, date, clock_in, total_break_minutes, clock_in_lat, clock_in_lng, clock_in_address, location_type) VALUES (%s, %s, %s, %s, 0, %s, %s, %s, %s)",
        (user["id"], user["name"], today, datetime.now(timezone.utc).isoformat(), lat, lng, address, location_type), last_id=True
    )

    return {"id": str(att_id), "user_id": str(user["id"]), "user_name": user["name"], "date": today, "clock_in": datetime.now(timezone.utc).isoformat(), "clock_out": None, "breaks": [], "total_break_minutes": 0, "clock_in_lat": lat, "clock_in_lng": lng, "clock_in_address": address, "location_type": location_type}

@attendance_router.post("/clock-out")
async def clock_out(request: Request):
    user = await get_current_user(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # Parse location from request body
    try:
        body = await request.json()
        lat = body.get("latitude")
        lng = body.get("longitude")
    except Exception:
        lat = None
        lng = None

    # Allow clock-out without GPS if user has approved WFH today
    wfh_today = await execute_query(
        "SELECT id FROM wfh_requests WHERE user_id = %s AND date = %s AND status = 'approved'",
        (user["id"], today), fetch_one=True
    )
    is_wfh = bool(wfh_today)

    if lat is None or lng is None:
        if not is_wfh:
            raise HTTPException(status_code=400, detail="Location is required for clock out.")
        lat = 0.0
        lng = 0.0

    attendance = await execute_query(
        "SELECT * FROM attendance WHERE user_id = %s AND date = %s AND clock_out IS NULL", (user["id"], today), fetch_one=True
    )
    if not attendance:
        raise HTTPException(status_code=400, detail="Not clocked in")

    clock_out_time = datetime.now(timezone.utc)

    # Reverse geocode (skip for WFH with no location)
    address = await reverse_geocode(lat, lng) if (lat != 0.0 or lng != 0.0) else "Work From Home"

    # End any active break
    active_break = await execute_query(
        "SELECT id, break_start FROM breaks WHERE attendance_id = %s AND break_end IS NULL", (attendance["id"],), fetch_one=True
    )
    if active_break:
        await execute_query("UPDATE breaks SET break_end = %s WHERE id = %s", (clock_out_time.isoformat(), active_break["id"]))

    # Calculate total break minutes
    breaks_list = await execute_query(
        "SELECT break_start, break_end FROM breaks WHERE attendance_id = %s", (attendance["id"],), fetch_all=True
    )
    total_break = 0
    for brk in (breaks_list or []):
        if brk["break_start"] and brk["break_end"]:
            s = datetime.fromisoformat(brk["break_start"])
            e = datetime.fromisoformat(brk["break_end"])
            total_break += (e - s).total_seconds() / 60

    clock_in_time = datetime.fromisoformat(attendance["clock_in"])
    total_time_minutes = (clock_out_time - clock_in_time).total_seconds() / 60
    # Only deduct break time that exceeds the allowed 30 min break
    excess_break = max(0, total_break - MAX_BREAK_MINUTES)
    working_minutes = total_time_minutes - excess_break
    working_hours = working_minutes / 60
    is_short_day = 1 if working_hours < REQUIRED_WORK_HOURS else 0

    await execute_query(
        "UPDATE attendance SET clock_out = %s, total_break_minutes = %s, working_hours = %s, is_short_day = %s, clock_out_lat = %s, clock_out_lng = %s, clock_out_address = %s WHERE id = %s",
        (clock_out_time.isoformat(), int(total_break), round(working_hours, 2), is_short_day, lat, lng, address, attendance["id"])
    )

    if is_short_day:
        await check_and_deduct_half_day_leave(user["id"])

    breaks_formatted = [{"start": b["break_start"], "end": b["break_end"]} for b in (breaks_list or [])]
    return {"id": str(attendance["id"]), "user_id": str(user["id"]), "user_name": user["name"], "date": today, "clock_in": attendance["clock_in"], "clock_out": clock_out_time.isoformat(), "breaks": breaks_formatted, "total_break_minutes": int(total_break), "working_hours": round(working_hours, 2), "is_short_day": bool(is_short_day), "clock_out_lat": lat, "clock_out_lng": lng, "clock_out_address": address}

async def check_and_deduct_half_day_leave(user_id: int):
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1).strftime("%Y-%m-%d")

    result = await execute_query(
        "SELECT COUNT(*) as cnt FROM attendance WHERE user_id = %s AND is_short_day = 1 AND date >= %s",
        (user_id, month_start), fetch_one=True
    )
    short_days_count = result["cnt"] if result else 0

    if short_days_count > 0 and short_days_count % SHORT_DAYS_FOR_HALF_LEAVE == 0:
        await execute_query(
            "UPDATE users SET casual_leave = casual_leave - 0.5, half_day_leave = half_day_leave + 0.5 WHERE id = %s", (user_id,)
        )
        await execute_query(
            "INSERT INTO leave_deductions (user_id, type, amount, reason, date) VALUES (%s, %s, %s, %s, %s)",
            (user_id, "half_day_short_work", 0.5, f"Auto-deducted for {SHORT_DAYS_FOR_HALF_LEAVE} short working days", now.isoformat())
        )

@attendance_router.post("/break/start")
async def start_break(request: Request):
    user = await get_current_user(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    try:
        body = await request.json()
        lat = body.get("latitude")
        lng = body.get("longitude")
    except Exception:
        lat = None
        lng = None

    if lat is None or lng is None:
        raise HTTPException(status_code=400, detail="Location is required for break.")

    attendance = await execute_query(
        "SELECT * FROM attendance WHERE user_id = %s AND date = %s AND clock_out IS NULL", (user["id"], today), fetch_one=True
    )
    if not attendance:
        raise HTTPException(status_code=400, detail="Not clocked in")

    active = await execute_query(
        "SELECT id FROM breaks WHERE attendance_id = %s AND break_end IS NULL", (attendance["id"],), fetch_one=True
    )
    if active:
        raise HTTPException(status_code=400, detail="Already on break")

    total_break_used = attendance.get("total_break_minutes", 0) or 0
    if total_break_used >= MAX_BREAK_MINUTES:
        raise HTTPException(status_code=400, detail=f"Break limit reached. Maximum {MAX_BREAK_MINUTES} minutes break allowed per day.")

    await execute_query(
        "INSERT INTO breaks (attendance_id, break_start, break_start_lat, break_start_lng) VALUES (%s, %s, %s, %s)",
        (attendance["id"], datetime.now(timezone.utc).isoformat(), lat, lng)
    )

    breaks_list = await execute_query("SELECT break_start, break_end FROM breaks WHERE attendance_id = %s", (attendance["id"],), fetch_all=True)
    breaks_formatted = [{"start": b["break_start"], "end": b["break_end"]} for b in (breaks_list or [])]

    return {"message": "Break started", "breaks": breaks_formatted, "remaining_break_minutes": MAX_BREAK_MINUTES - total_break_used, "max_break_minutes": MAX_BREAK_MINUTES}

@attendance_router.post("/break/end")
async def end_break(request: Request):
    user = await get_current_user(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    try:
        body = await request.json()
        lat = body.get("latitude")
        lng = body.get("longitude")
    except Exception:
        lat = None
        lng = None

    if lat is None or lng is None:
        raise HTTPException(status_code=400, detail="Location is required for ending break.")

    attendance = await execute_query(
        "SELECT * FROM attendance WHERE user_id = %s AND date = %s AND clock_out IS NULL", (user["id"], today), fetch_one=True
    )
    if not attendance:
        raise HTTPException(status_code=400, detail="Not clocked in")

    active = await execute_query(
        "SELECT id FROM breaks WHERE attendance_id = %s AND break_end IS NULL", (attendance["id"],), fetch_one=True
    )
    if not active:
        raise HTTPException(status_code=400, detail="Not on break")

    await execute_query("UPDATE breaks SET break_end = %s, break_end_lat = %s, break_end_lng = %s WHERE id = %s",
        (datetime.now(timezone.utc).isoformat(), lat, lng, active["id"]))

    breaks_list = await execute_query("SELECT break_start, break_end FROM breaks WHERE attendance_id = %s", (attendance["id"],), fetch_all=True)
    total_break = 0
    for brk in (breaks_list or []):
        if brk["break_start"] and brk["break_end"]:
            s = datetime.fromisoformat(brk["break_start"])
            e = datetime.fromisoformat(brk["break_end"])
            total_break += (e - s).total_seconds() / 60

    await execute_query("UPDATE attendance SET total_break_minutes = %s WHERE id = %s", (int(total_break), attendance["id"]))

    breaks_formatted = [{"start": b["break_start"], "end": b["break_end"]} for b in (breaks_list or [])]
    return {"message": "Break ended", "breaks": breaks_formatted, "total_break_minutes": int(total_break)}

@attendance_router.get("/status")
async def get_attendance_status(request: Request):
    user = await get_current_user(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    attendance = await execute_query(
        "SELECT * FROM attendance WHERE user_id = %s AND date = %s ORDER BY id DESC LIMIT 1",
        (user["id"], today), fetch_one=True
    )

    # Check if user has approved WFH today
    wfh_today = await execute_query(
        "SELECT id FROM wfh_requests WHERE user_id = %s AND date = %s AND status = 'approved'",
        (user["id"], today), fetch_one=True
    )
    has_wfh_today = bool(wfh_today)

    if not attendance:
        return {"clocked_in": False, "on_break": False, "attendance": None, "max_break_minutes": MAX_BREAK_MINUTES, "remaining_break_minutes": MAX_BREAK_MINUTES, "has_wfh_today": has_wfh_today}

    breaks_list = await execute_query("SELECT break_start, break_end FROM breaks WHERE attendance_id = %s", (attendance["id"],), fetch_all=True)
    breaks_formatted = [{"start": b["break_start"], "end": b["break_end"]} for b in (breaks_list or [])]

    on_break = any(b["end"] is None for b in breaks_formatted)
    total_break_used = attendance.get("total_break_minutes", 0) or 0
    remaining_break = max(0, MAX_BREAK_MINUTES - total_break_used)

    att_data = {
        "user_id": str(attendance["user_id"]),
        "user_name": attendance["user_name"],
        "date": attendance["date"],
        "clock_in": attendance["clock_in"],
        "clock_out": attendance["clock_out"],
        "breaks": breaks_formatted,
        "total_break_minutes": total_break_used,
        "working_hours": attendance.get("working_hours"),
        "is_short_day": bool(attendance.get("is_short_day"))
    }

    return {"clocked_in": attendance.get("clock_out") is None, "on_break": on_break, "attendance": att_data, "max_break_minutes": MAX_BREAK_MINUTES, "remaining_break_minutes": remaining_break, "has_wfh_today": has_wfh_today}

@attendance_router.get("/history")
async def get_attendance_history(request: Request, limit: int = 30):
    user = await get_current_user(request)
    records = await execute_query(
        "SELECT * FROM attendance WHERE user_id = %s ORDER BY date DESC LIMIT %s",
        (user["id"], limit), fetch_all=True
    )

    result = []
    for rec in (records or []):
        breaks_list = await execute_query("SELECT break_start, break_end FROM breaks WHERE attendance_id = %s", (rec["id"],), fetch_all=True)
        result.append({
            "user_id": str(rec["user_id"]),
            "user_name": rec["user_name"],
            "date": rec["date"],
            "clock_in": rec["clock_in"],
            "clock_out": rec["clock_out"],
            "breaks": [{"start": b["break_start"], "end": b["break_end"]} for b in (breaks_list or [])],
            "total_break_minutes": rec.get("total_break_minutes", 0),
            "working_hours": rec.get("working_hours"),
            "is_short_day": bool(rec.get("is_short_day"))
        })
    return result

@attendance_router.get("/working-hours-summary")
async def get_working_hours_summary(request: Request):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1).strftime("%Y-%m-%d")

    records = await execute_query(
        "SELECT working_hours, is_short_day FROM attendance WHERE user_id = %s AND date >= %s AND clock_out IS NOT NULL",
        (user["id"], month_start), fetch_all=True
    )

    total_working_hours = sum(r.get("working_hours", 0) or 0 for r in (records or []))
    short_days = sum(1 for r in (records or []) if r.get("is_short_day"))
    count = len(records or [])

    deductions = await execute_query(
        "SELECT amount FROM leave_deductions WHERE user_id = %s AND date >= %s",
        (user["id"], now.replace(day=1).isoformat()), fetch_all=True
    )
    total_deducted = sum(d.get("amount", 0) for d in (deductions or []))

    return {
        "total_working_days": count,
        "total_working_hours": round(total_working_hours, 2),
        "average_hours_per_day": round(total_working_hours / count, 2) if count else 0,
        "short_days_count": short_days,
        "half_days_deducted": total_deducted,
        "required_hours_per_day": REQUIRED_WORK_HOURS,
        "total_hours_per_day": TOTAL_WORK_HOURS,
        "short_days_for_half_leave": SHORT_DAYS_FOR_HALF_LEAVE
    }

@attendance_router.get("/my-shift")
async def get_my_shift(request: Request):
    user = await get_current_user(request)
    shift_str = user.get("shift") or ""
    info = parse_shift(shift_str)
    return {
        "shift": shift_str,
        "name": f"{info['start']} – {info['end']}",
        "start_time": info["start"],
        "end_time": info["end"],
        "is_set": bool(shift_str)
    }

# ============== LEAVE ROUTES ==============

@leave_router.get("/balance")
async def get_leave_balance(request: Request):
    user = await get_current_user(request)
    return {"casual": user.get("casual_leave"), "sick": user.get("sick_leave"), "loss_of_pay": user.get("loss_of_pay", 0), "wfh_limit": user.get("wfh_limit")}

@leave_router.post("/request")
async def create_leave_request(leave_data: LeaveRequest, request: Request):
    user = await get_current_user(request)

    start = datetime.strptime(leave_data.start_date, "%Y-%m-%d")
    end = datetime.strptime(leave_data.end_date, "%Y-%m-%d")

    # For half-day leave, start and end must be same date
    if leave_data.is_half_day:
        if leave_data.start_date != leave_data.end_date:
            raise HTTPException(status_code=400, detail="Half-day leave must be for a single date")
        working_days = 0.5
    else:
        working_days = 0
        current = start
        while current <= end:
            date_str = current.strftime("%Y-%m-%d")
            if not is_weekend(date_str) and not is_holiday(date_str):
                working_days += 1
            current += timedelta(days=1)

    if working_days == 0:
        raise HTTPException(status_code=400, detail="Selected dates only contain weekends/holidays")

    # Monthly leave limit check (1.5 days per month) - skip for loss_of_pay
    if leave_data.leave_type != "loss_of_pay":
        month_start = start.strftime("%Y-%m") + "-01"
        month_num = start.month
        year_num = start.year
        if month_num == 12:
            month_end = f"{year_num + 1}-01-01"
        else:
            month_end = f"{year_num}-{month_num + 1:02d}-01"

        used_this_month = await execute_query(
            "SELECT COALESCE(SUM(days), 0) as total FROM leave_requests WHERE user_id = %s AND start_date >= %s AND start_date < %s AND status IN ('pending', 'approved') AND leave_type != 'loss_of_pay'",
            (user["id"], month_start, month_end), fetch_one=True
        )
        used = float(used_this_month["total"]) if used_this_month else 0
        if used + working_days > 1.5:
            remaining = max(0, 1.5 - used)
            raise HTTPException(status_code=400, detail=f"Monthly leave limit is 1.5 days. Used: {used}, Remaining: {remaining}")

        # Check leave balance
        leave_field = f"{leave_data.leave_type}_leave"
        current_balance = user.get(leave_field, 0)
        if working_days > current_balance:
            raise HTTPException(status_code=400, detail=f"Insufficient {leave_data.leave_type} leave balance")

    leave_id = await execute_query(
        """INSERT INTO leave_requests (user_id, user_name, user_email, leave_type, start_date, end_date, days, reason, status, created_at, is_half_day)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'pending', %s, %s)""",
        (user["id"], user["name"], user.get("email", ""), leave_data.leave_type, leave_data.start_date, leave_data.end_date, working_days, leave_data.reason, datetime.now(timezone.utc).isoformat(), 1 if leave_data.is_half_day else 0),
        last_id=True
    )

    return {"id": str(leave_id), "user_id": str(user["id"]), "user_name": user["name"], "user_email": user.get("email", ""), "leave_type": leave_data.leave_type, "start_date": leave_data.start_date, "end_date": leave_data.end_date, "days": working_days, "reason": leave_data.reason, "status": "pending", "created_at": datetime.now(timezone.utc).isoformat(), "reviewed_by": None, "reviewed_at": None, "is_half_day": leave_data.is_half_day}

@leave_router.get("/my-requests")
async def get_my_leave_requests(request: Request):
    user = await get_current_user(request)
    rows = await execute_query(
        "SELECT * FROM leave_requests WHERE user_id = %s ORDER BY created_at DESC",
        (user["id"],), fetch_all=True
    )
    result = []
    for r in (rows or []):
        result.append({
            "id": str(r["id"]), "user_id": str(r["user_id"]), "user_name": r["user_name"], "user_email": r["user_email"],
            "leave_type": r["leave_type"], "start_date": r["start_date"], "end_date": r["end_date"],
            "days": r["days"], "reason": r["reason"], "status": r["status"],
            "created_at": r["created_at"], "reviewed_by": r["reviewed_by"], "reviewed_at": r["reviewed_at"],
            "is_half_day": bool(r.get("is_half_day", 0))
        })
    return result

@leave_router.get("/monthly-usage")
async def get_monthly_leave_usage(request: Request):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc)
    month_start = f"{now.year}-{now.month:02d}-01"
    if now.month == 12:
        month_end = f"{now.year + 1}-01-01"
    else:
        month_end = f"{now.year}-{now.month + 1:02d}-01"
    used = await execute_query(
        "SELECT COALESCE(SUM(days), 0) as total FROM leave_requests WHERE user_id = %s AND start_date >= %s AND start_date < %s AND status IN ('pending', 'approved') AND leave_type != 'loss_of_pay'",
        (user["id"], month_start, month_end), fetch_one=True
    )
    used_days = float(used["total"]) if used else 0
    return {"monthly_limit": 1.5, "used": used_days, "remaining": max(0, 1.5 - used_days)}

@leave_router.delete("/{leave_id}")
async def cancel_leave_request(leave_id: str, request: Request):
    user = await get_current_user(request)
    leave_req = await execute_query("SELECT * FROM leave_requests WHERE id = %s", (int(leave_id),), fetch_one=True)
    if not leave_req:
        raise HTTPException(status_code=404, detail="Leave request not found")
    if leave_req["user_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    if leave_req["status"] != "pending":
        raise HTTPException(status_code=400, detail="Can only cancel pending requests")
    await execute_query("DELETE FROM leave_requests WHERE id = %s", (int(leave_id),))
    return {"message": "Leave request cancelled"}

# ============== ADMIN ROUTES ==============

@admin_router.get("/employees")
async def get_all_employees(request: Request):
    await require_admin_or_manager(request)
    rows = await execute_query("SELECT id, email, name, role, department, position, avatar_url, created_at, casual_leave, sick_leave, loss_of_pay, permission_hours, half_day_leave, shift, basic_salary, employee_code, wfh_limit FROM users ORDER BY created_at DESC", fetch_all=True)
    result = []
    for emp in (rows or []):
        e = dict(emp)
        e["id"] = str(e.pop("id"))
        result.append(e)
    return result

@admin_router.get("/next-employee-code")
async def get_next_employee_code(request: Request):
    await require_admin(request)
    max_code = await execute_query("SELECT employee_code FROM users WHERE employee_code LIKE 'SC%' ORDER BY employee_code DESC LIMIT 1", fetch_one=True)
    next_num = 24001
    if max_code and max_code["employee_code"]:
        try:
            next_num = int(max_code["employee_code"][2:]) + 1
        except ValueError:
            pass
    return {"next_code": f"SC{next_num}"}

@admin_router.post("/employees")
async def create_employee(user_data: UserRegister, request: Request):
    await require_admin(request)
    email = user_data.email.lower()
    existing = await execute_query("SELECT id FROM users WHERE email = %s", (email,), fetch_one=True)
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")

    role = user_data.role if user_data.role in ("employee", "manager") else "employee"
    hashed = hash_password(user_data.password)

    # Generate next employee code (SC24001, SC24002, ...) or use provided
    if user_data.employee_code and user_data.employee_code.strip():
        custom_code = user_data.employee_code.strip().upper()
        dup = await execute_query("SELECT id FROM users WHERE employee_code = %s", (custom_code,), fetch_one=True)
        if dup:
            raise HTTPException(status_code=400, detail=f"Employee code '{custom_code}' already exists")
        employee_code = custom_code
    else:
        max_code = await execute_query("SELECT employee_code FROM users WHERE employee_code LIKE 'SC%' ORDER BY employee_code DESC LIMIT 1", fetch_one=True)
        next_num = 24001
        if max_code and max_code["employee_code"]:
            try:
                next_num = int(max_code["employee_code"][2:]) + 1
            except ValueError:
                pass
        employee_code = f"SC{next_num}"

    user_id = await execute_query(
        """INSERT INTO users (email, password_hash, name, role, department, position, avatar_url, created_at, casual_leave, sick_leave, loss_of_pay, permission_hours, half_day_leave, shift, employee_code, wfh_limit)
           VALUES (%s, %s, %s, %s, %s, %s, '', %s, 12, 3, 0, %s, 0, '', %s, 4)""",
        (email, hashed, user_data.name, role, user_data.department, user_data.position, datetime.now(timezone.utc).isoformat(), MONTHLY_PERMISSION_HOURS, employee_code),
        last_id=True
    )

    return {"id": str(user_id), "email": email, "name": user_data.name, "role": role, "department": user_data.department, "position": user_data.position, "avatar_url": "", "casual_leave": None, "sick_leave": None, "loss_of_pay": 0, "permission_hours": MONTHLY_PERMISSION_HOURS, "half_day_leave": 0, "shift": "", "employee_code": employee_code, "wfh_limit": None}

@admin_router.put("/employees/{employee_id}")
async def update_employee(employee_id: str, update_data: EmployeeUpdate, request: Request):
    await require_admin(request)
    # Use model_fields_set so explicitly-sent null values are included (to clear fields)
    update_dict = {k: v for k, v in update_data.model_dump().items() if k in update_data.model_fields_set}
    if not update_dict:
        raise HTTPException(status_code=400, detail="No update data provided")

    # Validate role if being changed
    if "role" in update_dict and update_dict["role"] not in ("employee", "manager", "admin"):
        raise HTTPException(status_code=400, detail="Invalid role")

    # Validate employee_code uniqueness if being changed
    if "employee_code" in update_dict and update_dict["employee_code"]:
        code = update_dict["employee_code"].strip().upper()
        update_dict["employee_code"] = code
        dup = await execute_query("SELECT id FROM users WHERE employee_code = %s AND id != %s", (code, int(employee_id)), fetch_one=True)
        if dup:
            raise HTTPException(status_code=400, detail=f"Employee code '{code}' already assigned to another employee")

    set_clause = ", ".join([f"{k} = %s" for k in update_dict])
    values = list(update_dict.values()) + [int(employee_id)]
    await execute_query(f"UPDATE users SET {set_clause} WHERE id = %s", tuple(values))

    emp = await execute_query("SELECT id, email, name, role, department, position, avatar_url, casual_leave, sick_leave, loss_of_pay, permission_hours, half_day_leave, shift, basic_salary, employee_code, wfh_limit FROM users WHERE id = %s", (int(employee_id),), fetch_one=True)
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    emp["id"] = str(emp.pop("id"))
    return emp

@admin_router.delete("/employees/{employee_id}")
async def delete_employee(employee_id: str, request: Request):
    await require_admin(request)
    result = await execute_query("DELETE FROM users WHERE id = %s AND role != 'admin'", (int(employee_id),))
    if result == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    return {"message": "Employee deleted"}

@admin_router.post("/employees/{employee_id}/reset-password")
async def reset_employee_password(employee_id: str, data: PasswordReset, request: Request):
    await require_admin(request)
    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    hashed = hash_password(data.new_password)
    result = await execute_query("UPDATE users SET password_hash = %s WHERE id = %s AND role != 'admin'", (hashed, int(employee_id)))
    if result == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    return {"message": "Password reset successfully"}

@admin_router.post("/employees/{employee_id}/avatar")
async def upload_employee_avatar(employee_id: str, request: Request, file: UploadFile = File(...)):
    await require_admin(request)
    
    # Validate file type
    allowed_types = ["image/jpeg", "image/png", "image/webp", "image/gif"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Only JPEG, PNG, WebP, GIF images allowed")
    
    # Validate file size (max 5MB)
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File size must be under 5MB")
    
    # Check employee exists
    emp = await execute_query("SELECT id FROM users WHERE id = %s", (int(employee_id),), fetch_one=True)
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Save to MySQL media table
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "jpg"
    filename = f"avatar_{employee_id}_{_uuid.uuid4().hex[:8]}.{ext}"

    try:
        media_id = await _save_image_db(contents, filename)
    except Exception as e:
        logging.error(f"Avatar upload failed: {e}")
        raise HTTPException(status_code=500, detail="Avatar upload failed")

    avatar_url = f"/api/admin/avatars/{media_id}"
    await execute_query("UPDATE users SET avatar_url = %s WHERE id = %s", (avatar_url, int(employee_id)))
    return {"message": "Avatar uploaded", "avatar_url": avatar_url}

@admin_router.get("/avatars/{media_id}")
async def serve_avatar(media_id: str):
    """Serve avatar images from MySQL media table."""
    try:
        data, content_type = await _get_image_db(media_id)
        return Response(content=data, media_type=content_type, headers={
            "Cache-Control": "public, max-age=86400"
        })
    except Exception as e:
        logging.error(f"Avatar fetch failed for {media_id}: {e}")
        raise HTTPException(status_code=404, detail="Avatar not found")

# ── Leave Balance ─────────────────────────────────────────────────────────────
class LeaveBalanceUpdate(BaseModel):
    casual_leave: Optional[float] = None
    sick_leave: Optional[float] = None
    loss_of_pay: Optional[float] = None
    permission_hours: Optional[float] = None
    wfh_limit: Optional[int] = None

@admin_router.put("/employees/{employee_id}/leave-balance")
async def update_leave_balance(employee_id: str, data: LeaveBalanceUpdate, request: Request):
    await require_admin(request)
    emp = await execute_query("SELECT id FROM users WHERE id = %s", (int(employee_id),), fetch_one=True)
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    fields, values = [], []
    # All fields: sent as null = clear to null, not sent = skip
    for field in ["casual_leave", "sick_leave", "loss_of_pay", "permission_hours", "wfh_limit"]:
        if field in data.model_fields_set:
            fields.append(f"{field} = %s"); values.append(getattr(data, field))
    if not fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    values.append(int(employee_id))
    await execute_query(f"UPDATE users SET {', '.join(fields)} WHERE id = %s", tuple(values))
    return {"message": "Leave balance updated successfully"}

@admin_router.get("/employees/{employee_id}/leave-balance")
async def get_employee_leave_balance(employee_id: str, request: Request):
    await require_admin(request)
    emp = await execute_query(
        "SELECT casual_leave, sick_leave, loss_of_pay, permission_hours, wfh_limit FROM users WHERE id = %s",
        (int(employee_id),), fetch_one=True
    )
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    return emp

# ── Org Chart ─────────────────────────────────────────────────────────────────
class OrgNodeCreate(BaseModel):
    employee_name: str
    job_title: Optional[str] = ""
    image_url: Optional[str] = ""
    description: Optional[str] = ""
    parent_id: Optional[int] = None
    level_num: Optional[int] = 0
    sort_order: Optional[int] = 0

class OrgNodeUpdate(BaseModel):
    employee_name: Optional[str] = None
    job_title: Optional[str] = None
    image_url: Optional[str] = None
    description: Optional[str] = None
    parent_id: Optional[int] = None
    level_num: Optional[int] = None
    sort_order: Optional[int] = None

@admin_router.get("/org-chart")
async def get_org_chart(request: Request):
    await get_current_user(request)  # any logged in user can view
    rows = await execute_query(
        "SELECT * FROM org_chart ORDER BY sort_order ASC, id ASC",
        fetch_all=True
    )
    return rows or []

# ── Org Levels (custom label per depth level) ─────────────────────────────────
@app.get("/api/org-levels")
async def get_org_levels(request: Request):
    await get_current_user(request)
    rows = await execute_query("SELECT level_num, label FROM org_levels ORDER BY level_num ASC", fetch_all=True)
    return rows or []

@admin_router.put("/org-levels")
async def save_org_levels(request: Request):
    await require_admin(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")
    if not isinstance(body, list):
        raise HTTPException(status_code=400, detail="Expected a list of {level_num, label}")
    now = datetime.now(timezone.utc).isoformat()
    for item in body:
        lnum = int(item.get("level_num", 0))
        label = str(item.get("label", f"Level {lnum + 1}")).strip() or f"Level {lnum + 1}"
        await execute_query(
            "INSERT INTO org_levels (level_num, label, updated_at) VALUES (%s, %s, %s) ON DUPLICATE KEY UPDATE label=%s, updated_at=%s",
            (lnum, label, now, label, now)
        )
    return {"message": "Org levels saved"}

@admin_router.post("/org-chart")
async def create_org_node(data: OrgNodeCreate, request: Request):
    await require_admin(request)
    from datetime import datetime, timezone as tz
    now = datetime.now(tz.utc).isoformat()
    result = await execute_query(
        "INSERT INTO org_chart (parent_id, employee_name, job_title, image_url, description, level_num, sort_order, created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
        (data.parent_id, data.employee_name, data.job_title, data.image_url, data.description, data.level_num or 0, data.sort_order, now),
        last_id=True
    )
    return {"id": result, "message": "Node created"}

@admin_router.put("/org-chart/{node_id}")
async def update_org_node(node_id: int, data: OrgNodeUpdate, request: Request):
    await require_admin(request)
    fields, values = [], []
    if data.employee_name is not None: fields.append("employee_name=%s"); values.append(data.employee_name)
    if data.job_title is not None: fields.append("job_title=%s"); values.append(data.job_title)
    if data.image_url is not None: fields.append("image_url=%s"); values.append(data.image_url)
    if data.description is not None: fields.append("description=%s"); values.append(data.description)
    if "parent_id" in data.model_fields_set: fields.append("parent_id=%s"); values.append(data.parent_id)
    if data.level_num is not None: fields.append("level_num=%s"); values.append(data.level_num)
    if data.sort_order is not None: fields.append("sort_order=%s"); values.append(data.sort_order)
    if not fields:
        raise HTTPException(status_code=400, detail="Nothing to update")
    values.append(node_id)
    await execute_query(f"UPDATE org_chart SET {', '.join(fields)} WHERE id=%s", tuple(values))
    return {"message": "Node updated"}

@admin_router.delete("/org-chart/{node_id}")
async def delete_org_node(node_id: int, request: Request):
    await require_admin(request)
    # Re-parent children to deleted node's parent
    node = await execute_query("SELECT parent_id FROM org_chart WHERE id=%s", (node_id,), fetch_one=True)
    if node:
        await execute_query("UPDATE org_chart SET parent_id=%s WHERE parent_id=%s", (node["parent_id"], node_id))
    await execute_query("DELETE FROM org_chart WHERE id=%s", (node_id,))
    return {"message": "Node deleted"}

@admin_router.post("/org-chart/{node_id}/image")
async def upload_org_node_image(node_id: int, file: UploadFile = File(...), request: Request = None):
    await require_admin(request)
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Image too large (max 5MB)")
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "jpg"
    filename = f"org_{node_id}_{_uuid.uuid4().hex[:8]}.{ext}"
    try:
        media_id = await _save_image_db(contents, filename)
    except Exception as e:
        logging.error(f"Org node image upload failed for node {node_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Image upload failed: {str(e)}")
    image_url = f"/api/admin/avatars/{media_id}"
    await execute_query("UPDATE org_chart SET image_url=%s WHERE id=%s", (image_url, node_id))
    return {"image_url": image_url}

# ── Role Permissions ──────────────────────────────────────────────────────────
@admin_router.get("/role-permissions")
async def get_role_permissions(request: Request):
    await require_admin(request)
    rows = await execute_query("SELECT role, feature_key, enabled FROM role_permissions ORDER BY role, feature_key", fetch_all=True)
    result = {}
    for r in (rows or []):
        role = r["role"]
        if role not in result:
            result[role] = {}
        result[role][r["feature_key"]] = bool(r["enabled"])
    return result

@admin_router.put("/role-permissions")
async def update_role_permissions(request: Request):
    await require_admin(request)
    body = await request.json()  # { role: { feature_key: bool, ... }, ... }
    for role, features in body.items():
        if role not in ("manager", "employee"):
            continue
        for feature_key, enabled in features.items():
            await execute_query(
                "INSERT INTO role_permissions (role, feature_key, enabled) VALUES (%s,%s,%s) ON DUPLICATE KEY UPDATE enabled=%s",
                (role, feature_key, int(bool(enabled)), int(bool(enabled)))
            )
    return {"message": "Permissions updated"}

# All roles can read their own permissions
@app.get("/api/my-permissions")
async def get_my_permissions(request: Request):
    user = await get_current_user(request)
    role = user.get("role", "employee")
    if role == "admin":
        return {"role": "admin", "permissions": {}}  # admin has all access
    rows = await execute_query(
        "SELECT feature_key, enabled FROM role_permissions WHERE role=%s",
        (role,), fetch_all=True
    )
    perms = {}
    for r in (rows or []):
        perms[r["feature_key"]] = bool(r["enabled"])
    return {"role": role, "permissions": perms}

@admin_router.get("/leave-requests")
async def get_all_leave_requests(request: Request, status: Optional[str] = None):
    user = await require_admin_or_manager(request)
    query = "SELECT * FROM leave_requests"
    args = []
    conditions = []
    if status:
        conditions.append("status = %s")
        args.append(status)
    # Managers only see leave requests from their department employees
    if user["role"] == "manager":
        conditions.append("user_id IN (SELECT id FROM users WHERE department = %s)")
        args.append(user.get("department", ""))
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY created_at DESC"
    rows = await execute_query(query, tuple(args) if args else None, fetch_all=True)
    result = []
    for r in (rows or []):
        d = dict(r)
        d["id"] = str(d.pop("id"))
        d["user_id"] = str(d["user_id"])
        result.append(d)
    return result

@admin_router.put("/leave-requests/{leave_id}")
async def review_leave_request(leave_id: str, request: Request, action: str):
    reviewer = await require_admin_or_manager(request)
    if action not in ("approve", "reject"):
        raise HTTPException(status_code=400, detail="Invalid action")

    leave_req = await execute_query("SELECT * FROM leave_requests WHERE id = %s", (int(leave_id),), fetch_one=True)
    if not leave_req:
        raise HTTPException(status_code=404, detail="Leave request not found")
    if leave_req["status"] != "pending":
        raise HTTPException(status_code=400, detail="Leave request already processed")

    new_status = "approved" if action == "approve" else "rejected"
    await execute_query(
        "UPDATE leave_requests SET status = %s, reviewed_by = %s, reviewed_at = %s WHERE id = %s",
        (new_status, reviewer["name"], datetime.now(timezone.utc).isoformat(), int(leave_id))
    )

    if action == "approve":
        if leave_req["leave_type"] == "loss_of_pay":
            await execute_query("UPDATE users SET loss_of_pay = loss_of_pay + %s WHERE id = %s", (leave_req["days"], leave_req["user_id"]))
        else:
            leave_field = f"{leave_req['leave_type']}_leave"
            await execute_query(f"UPDATE users SET {leave_field} = {leave_field} - %s WHERE id = %s", (leave_req["days"], leave_req["user_id"]))

    return {"message": f"Leave request {new_status}"}

@admin_router.get("/attendance")
async def get_all_attendance(request: Request, date: Optional[str] = None):
    await require_admin_or_manager(request)
    query = "SELECT * FROM attendance"
    args = []
    if date:
        query += " WHERE date = %s"
        args.append(date)
    query += " ORDER BY date DESC LIMIT 1000"
    rows = await execute_query(query, tuple(args) if args else None, fetch_all=True)

    result = []
    for rec in (rows or []):
        breaks_list = await execute_query("SELECT break_start, break_end, break_start_lat, break_start_lng FROM breaks WHERE attendance_id = %s", (rec["id"],), fetch_all=True)
        # Check if any break was outside geofence
        break_outside = False
        if breaks_list:
            office = await get_office_settings()
            for brk in breaks_list:
                if brk.get("break_start_lat") and brk.get("break_start_lng"):
                    dist = haversine_km(brk["break_start_lat"], brk["break_start_lng"], office["latitude"], office["longitude"])
                    if dist > office["radius_km"]:
                        break_outside = True
                        break
        result.append({
            "user_id": str(rec["user_id"]), "user_name": rec["user_name"], "date": rec["date"],
            "clock_in": rec["clock_in"], "clock_out": rec["clock_out"],
            "breaks": [{"start": b["break_start"], "end": b["break_end"]} for b in (breaks_list or [])],
            "total_break_minutes": rec.get("total_break_minutes", 0),
            "working_hours": rec.get("working_hours"), "is_short_day": bool(rec.get("is_short_day")),
            "clock_in_lat": rec.get("clock_in_lat"), "clock_in_lng": rec.get("clock_in_lng"),
            "clock_in_address": rec.get("clock_in_address"),
            "clock_out_address": rec.get("clock_out_address"),
            "location_type": rec.get("location_type"),
            "break_outside_geofence": break_outside
        })
    return result

@admin_router.get("/analytics")
async def get_analytics(request: Request):
    await require_admin_or_manager(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    total_emp = await execute_query("SELECT COUNT(*) as cnt FROM users WHERE role IN ('employee', 'manager')", fetch_one=True)
    present = await execute_query("SELECT COUNT(*) as cnt FROM attendance WHERE date = %s", (today,), fetch_one=True)
    pending = await execute_query("SELECT COUNT(*) as cnt FROM leave_requests WHERE status = 'pending'", fetch_one=True)

    # On break
    today_att = await execute_query("SELECT id FROM attendance WHERE date = %s AND clock_out IS NULL", (today,), fetch_all=True)
    on_break = 0
    for att in (today_att or []):
        ab = await execute_query("SELECT id FROM breaks WHERE attendance_id = %s AND break_end IS NULL", (att["id"],), fetch_one=True)
        if ab:
            on_break += 1

    # Department breakdown
    dept = await execute_query("SELECT department, COUNT(*) as count FROM users WHERE role IN ('employee', 'manager') GROUP BY department", fetch_all=True)

    total = total_emp["cnt"] if total_emp else 0
    return {
        "total_employees": total,
        "present_today": present["cnt"] if present else 0,
        "absent_today": total - (present["cnt"] if present else 0),
        "pending_leaves": pending["cnt"] if pending else 0,
        "on_break": on_break,
        "department_breakdown": [{"department": d["department"], "count": d["count"]} for d in (dept or [])]
    }

# ============== PERMISSION ROUTES ==============

@permission_router.get("/balance")
async def get_permission_balance(request: Request):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1).strftime("%Y-%m-%d")

    rows = await execute_query(
        "SELECT duration_minutes FROM permissions WHERE user_id = %s AND date >= %s AND status IN ('approved', 'pending')",
        (user["id"], month_start), fetch_all=True
    )
    used_minutes = sum(r["duration_minutes"] for r in (rows or []))
    return {
        "monthly_allowance_hours": MONTHLY_PERMISSION_HOURS, "used_minutes": used_minutes,
        "used_hours": used_minutes / 60, "remaining_minutes": (MONTHLY_PERMISSION_HOURS * 60) - used_minutes,
        "remaining_hours": MONTHLY_PERMISSION_HOURS - (used_minutes / 60), "max_per_use_minutes": MAX_PERMISSION_PER_USE * 60
    }

@permission_router.post("/request")
async def request_permission(perm_data: PermissionRequest, request: Request):
    user = await get_current_user(request)
    if perm_data.duration_minutes > MAX_PERMISSION_PER_USE * 60:
        raise HTTPException(status_code=400, detail=f"Maximum permission duration is {MAX_PERMISSION_PER_USE} hour(s)")
    if perm_data.duration_minutes <= 0:
        raise HTTPException(status_code=400, detail="Invalid duration")

    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1).strftime("%Y-%m-%d")
    rows = await execute_query(
        "SELECT duration_minutes FROM permissions WHERE user_id = %s AND date >= %s AND status IN ('approved', 'pending')",
        (user["id"], month_start), fetch_all=True
    )
    used_minutes = sum(r["duration_minutes"] for r in (rows or []))
    remaining = (MONTHLY_PERMISSION_HOURS * 60) - used_minutes
    if perm_data.duration_minutes > remaining:
        raise HTTPException(status_code=400, detail=f"Insufficient permission balance. Remaining: {remaining} minutes")

    perm_id = await execute_query(
        "INSERT INTO permissions (user_id, user_name, user_email, duration_minutes, reason, date, status, created_at) VALUES (%s, %s, %s, %s, %s, %s, 'pending', %s)",
        (user["id"], user["name"], user.get("email", ""), perm_data.duration_minutes, perm_data.reason, perm_data.date, now.isoformat()),
        last_id=True
    )
    return {"id": str(perm_id), "user_id": str(user["id"]), "user_name": user["name"], "user_email": user.get("email", ""), "duration_minutes": perm_data.duration_minutes, "reason": perm_data.reason, "date": perm_data.date, "status": "pending", "created_at": now.isoformat(), "reviewed_by": None, "reviewed_at": None}

@permission_router.get("/my-requests")
async def get_my_permission_requests(request: Request):
    user = await get_current_user(request)
    rows = await execute_query("SELECT * FROM permissions WHERE user_id = %s ORDER BY created_at DESC", (user["id"],), fetch_all=True)
    result = []
    for r in (rows or []):
        d = dict(r)
        d["id"] = str(d.pop("id"))
        d["user_id"] = str(d["user_id"])
        result.append(d)
    return result

@permission_router.delete("/{permission_id}")
async def cancel_permission(permission_id: str, request: Request):
    user = await get_current_user(request)
    perm = await execute_query("SELECT * FROM permissions WHERE id = %s", (int(permission_id),), fetch_one=True)
    if not perm:
        raise HTTPException(status_code=404, detail="Permission request not found")
    if perm["user_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    if perm["status"] != "pending":
        raise HTTPException(status_code=400, detail="Can only cancel pending requests")
    await execute_query("DELETE FROM permissions WHERE id = %s", (int(permission_id),))
    return {"message": "Permission request cancelled"}

@admin_router.get("/permissions")
async def get_all_permissions(request: Request, status: Optional[str] = None):
    user = await require_admin_or_manager(request)
    query = "SELECT * FROM permissions"
    args = []
    conditions = []
    if status:
        conditions.append("status = %s")
        args.append(status)
    if user["role"] == "manager":
        conditions.append("user_id IN (SELECT id FROM users WHERE department = %s)")
        args.append(user.get("department", ""))
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY created_at DESC"
    rows = await execute_query(query, tuple(args) if args else None, fetch_all=True)
    result = []
    for r in (rows or []):
        d = dict(r)
        d["id"] = str(d.pop("id"))
        d["user_id"] = str(d["user_id"])
        result.append(d)
    return result

@admin_router.put("/permissions/{permission_id}")
async def review_permission(permission_id: str, request: Request, action: str):
    reviewer = await require_admin_or_manager(request)
    if action not in ("approve", "reject"):
        raise HTTPException(status_code=400, detail="Invalid action")
    perm = await execute_query("SELECT * FROM permissions WHERE id = %s", (int(permission_id),), fetch_one=True)
    if not perm:
        raise HTTPException(status_code=404, detail="Permission request not found")
    if perm["status"] != "pending":
        raise HTTPException(status_code=400, detail="Permission already processed")
    new_status = "approved" if action == "approve" else "rejected"
    await execute_query(
        "UPDATE permissions SET status = %s, reviewed_by = %s, reviewed_at = %s WHERE id = %s",
        (new_status, reviewer["name"], datetime.now(timezone.utc).isoformat(), int(permission_id))
    )
    return {"message": f"Permission request {new_status}"}

# ============== PAYSLIP ROUTES ==============

@payslip_router.get("/my-payslips")
async def get_my_payslips(request: Request):
    user = await get_current_user(request)
    rows = await execute_query("SELECT * FROM payslips WHERE employee_id = %s ORDER BY created_at DESC", (user["id"],), fetch_all=True)
    result = []
    for ps in (rows or []):
        d = dict(ps)
        d["id"] = str(d.pop("id"))
        d["employee_id"] = str(d["employee_id"])
        if isinstance(d.get("deduction_details"), str):
            d["deduction_details"] = json.loads(d["deduction_details"])
        result.append(d)
    return result

@payslip_router.get("/download/{payslip_id}")
async def download_payslip(payslip_id: str, request: Request):
    user = await get_current_user(request)
    payslip = await execute_query("SELECT * FROM payslips WHERE id = %s", (int(payslip_id),), fetch_one=True)
    if not payslip:
        raise HTTPException(status_code=404, detail="Payslip not found")
    if payslip["employee_id"] != user["id"] and user.get("role") not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Not authorized")

    ps = dict(payslip)
    if isinstance(ps.get("deduction_details"), str):
        ps["deduction_details"] = json.loads(ps["deduction_details"])
    # Fetch employee_code for the payslip
    emp = await execute_query("SELECT employee_code FROM users WHERE id = %s", (ps["employee_id"],), fetch_one=True)
    ps["employee_code"] = emp["employee_code"] if emp else ""
    pdf_buffer = generate_payslip_pdf(ps)
    return StreamingResponse(pdf_buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=payslip_{ps['month']}_{ps['year']}.pdf"})

def generate_payslip_pdf(payslip: dict) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=25, bottomMargin=25, leftMargin=35, rightMargin=35)
    elements = []
    styles = getSampleStyleSheet()
    page_width = A4[0] - 70

    # Colors matching the reference
    dark_text = colors.HexColor('#1a1a2e')
    medium_text = colors.HexColor('#4a4a5a')
    light_text = colors.HexColor('#8a8a9a')
    header_bg = colors.HexColor('#f7f7fa')
    border_color = colors.HexColor('#e0e0e8')
    green_bg = colors.HexColor('#e8f5e9')
    green_border = colors.HexColor('#4caf50')
    white = colors.white

    # ============ HEADER: Logo + Company + Net Pay ============
    logo_path = os.path.join(os.path.dirname(__file__), "uploads", "sparkcurv-logo.jpg")

    # Company info (left side)
    company_name_style = ParagraphStyle('CN', fontSize=13, fontName='Helvetica-Bold', textColor=dark_text, leading=16)
    company_addr_style = ParagraphStyle('CA', fontSize=8, textColor=light_text, leading=11)

    company_block = []
    company_block.append(Paragraph("SPARKCURV TECHNOLOGIES PVT LIMITED", company_name_style))
    company_block.append(Spacer(1, 3))
    company_block.append(Paragraph("64/3 Thompson Street, Palace Road, Nagercoil, Tamil Nadu 629001", company_addr_style))

    # Right side - Payslip month only
    month_label_style = ParagraphStyle('ML', fontSize=8, textColor=light_text, alignment=2)
    month_value_style = ParagraphStyle('MV', fontSize=11, fontName='Helvetica-Bold', textColor=dark_text, alignment=2)

    right_block = []
    right_block.append(Paragraph("Payslip For the Month", month_label_style))
    right_block.append(Paragraph(f"{payslip['month_name']} {payslip['year']}", month_value_style))

    # Build header with logo
    if os.path.exists(logo_path):
        try:
            logo = RLImage(logo_path, width=48, height=48)
            logo_cell = Table([[logo]], colWidths=[55])
            logo_cell.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
        except Exception:
            logo_cell = Paragraph("", styles['Normal'])
    else:
        logo_cell = Paragraph("", styles['Normal'])

    left_table = Table([[logo_cell, company_block]], colWidths=[58, page_width * 0.45])
    left_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('LEFTPADDING', (0, 0), (-1, -1), 0), ('RIGHTPADDING', (0, 0), (-1, -1), 0)]))

    right_table = Table([[right_block]], colWidths=[page_width * 0.45])
    right_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP'), ('ALIGN', (0, 0), (-1, -1), 'RIGHT')]))

    header = Table([[left_table, right_table]], colWidths=[page_width * 0.55, page_width * 0.45])
    header.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    elements.append(header)
    elements.append(Spacer(1, 12))
    elements.append(HRFlowable(width="100%", thickness=1, color=border_color))
    elements.append(Spacer(1, 12))

    # ============ EMPLOYEE SUMMARY ============
    section_title = ParagraphStyle('ST', fontSize=10, fontName='Helvetica-Bold', textColor=dark_text, spaceBefore=4, spaceAfter=8)
    label_s = ParagraphStyle('LS', fontSize=8, textColor=light_text, leading=12)
    value_s = ParagraphStyle('VS', fontSize=8, textColor=dark_text, fontName='Helvetica-Bold', leading=12)

    elements.append(Paragraph("EMPLOYEE SUMMARY", section_title))

    # Calculate paid days and LOP days
    basic_salary = payslip.get('basic_salary', 0) or 0
    net_pay = payslip.get('net_pay', 0) or 0
    lop_days = 0
    for ded in payslip.get('deduction_details', []):
        desc = ded.get('description', '')
        if 'Loss of Pay' in desc:
            import re
            match = re.search(r'\((\d+)\s*days?\)', desc)
            if match:
                lop_days = int(match.group(1))
    paid_days = WORKING_DAYS_PER_MONTH - lop_days

    pay_month = payslip['month'] + 1
    pay_year = payslip['year']
    if pay_month > 12:
        pay_month = 1
        pay_year += 1
    pay_date = f"{5:02d}/{pay_month:02d}/{pay_year}"

    emp_code = payslip.get('employee_code', '') or str(payslip.get('employee_id', ''))
    col_w = page_width / 4

    emp_rows = [
        [Paragraph("Employee Name", label_s), Paragraph(f": {payslip.get('employee_name', '')}", value_s),
         Paragraph("Pay Period", label_s), Paragraph(f": {payslip['month_name']} {payslip['year']}", value_s)],
        [Paragraph("Employee ID", label_s), Paragraph(f": {emp_code}", value_s),
         Paragraph("Pay Date", label_s), Paragraph(f": {pay_date}", value_s)],
        [Paragraph("Department", label_s), Paragraph(f": {payslip.get('department', 'N/A')}", value_s),
         Paragraph("Paid Days", label_s), Paragraph(f": {paid_days}", value_s)],
        [Paragraph("Designation", label_s), Paragraph(f": {payslip.get('position', 'N/A')}", value_s),
         Paragraph("LOP Days", label_s), Paragraph(f": {lop_days}", value_s)],
    ]

    emp_table = Table(emp_rows, colWidths=[col_w * 0.8, col_w * 1.2, col_w * 0.7, col_w * 1.3])
    emp_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(emp_table)
    elements.append(Spacer(1, 12))
    elements.append(HRFlowable(width="100%", thickness=1, color=border_color))
    elements.append(Spacer(1, 12))

    # ============ EARNINGS & DEDUCTIONS SIDE BY SIDE ============
    # Salary breakdown
    basic = round(basic_salary * 0.50, 2)
    hra = round(basic_salary * 0.20, 2)
    medical = round(basic_salary * 0.045, 2)
    conveyance = round(basic_salary * 0.06, 2)
    special = round(basic_salary - basic - hra - medical - conveyance, 2)
    gross = basic_salary

    h_style = ParagraphStyle('HS', fontSize=8, fontName='Helvetica-Bold', textColor=dark_text)
    h_r_style = ParagraphStyle('HRS', fontSize=8, fontName='Helvetica-Bold', textColor=dark_text, alignment=2)
    r_style = ParagraphStyle('RS', fontSize=8, textColor=medium_text)
    r_r_style = ParagraphStyle('RRS', fontSize=8, textColor=medium_text, alignment=2)
    t_style = ParagraphStyle('TS', fontSize=8, fontName='Helvetica-Bold', textColor=dark_text)
    t_r_style = ParagraphStyle('TRS', fontSize=8, fontName='Helvetica-Bold', textColor=dark_text, alignment=2)

    half_w = page_width / 2 - 6

    # Earnings table
    earn_rows = [
        [Paragraph("EARNINGS", h_style), Paragraph("AMOUNT", h_r_style)],
        [Paragraph("Basic", r_style), Paragraph(f"{basic:,.2f}", r_r_style)],
        [Paragraph("House Rent Allowance", r_style), Paragraph(f"{hra:,.2f}", r_r_style)],
        [Paragraph("Medical Allowance", r_style), Paragraph(f"{medical:,.2f}", r_r_style)],
        [Paragraph("Conveyance Allowance", r_style), Paragraph(f"{conveyance:,.2f}", r_r_style)],
        [Paragraph("Special Allowance", r_style), Paragraph(f"{special:,.2f}", r_r_style)],
        [Paragraph("Gross Earnings", t_style), Paragraph(f"{gross:,.2f}", t_r_style)],
    ]

    e_table = Table(earn_rows, colWidths=[half_w * 0.65, half_w * 0.35])
    e_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), header_bg),
        ('LINEBELOW', (0, 0), (-1, 0), 1, border_color),
        ('LINEBELOW', (0, 1), (-1, -2), 0.5, border_color),
        ('LINEABOVE', (0, -1), (-1, -1), 1, border_color),
        ('BACKGROUND', (0, -1), (-1, -1), header_bg),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))

    # Deductions table
    total_deductions = payslip.get('total_deductions', 0) or 0
    deduction_details = payslip.get('deduction_details', [])

    ded_rows = [
        [Paragraph("DEDUCTIONS", h_style), Paragraph("AMOUNT", h_r_style)],
    ]
    if deduction_details:
        for ded in deduction_details:
            ded_rows.append([
                Paragraph(ded.get('description', ''), r_style),
                Paragraph(f"{ded.get('amount', 0):,.2f}", r_r_style)
            ])
    else:
        ded_rows.append([Paragraph("No Deductions", r_style), Paragraph("0.00", r_r_style)])

    # Pad rows to match earnings height
    while len(ded_rows) < len(earn_rows) - 1:
        ded_rows.append([Paragraph("", r_style), Paragraph("", r_r_style)])

    ded_rows.append([Paragraph("Total Deductions", t_style), Paragraph(f"{total_deductions:,.2f}", t_r_style)])

    d_table = Table(ded_rows, colWidths=[half_w * 0.65, half_w * 0.35])
    d_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), header_bg),
        ('LINEBELOW', (0, 0), (-1, 0), 1, border_color),
        ('LINEBELOW', (0, 1), (-1, -2), 0.5, border_color),
        ('LINEABOVE', (0, -1), (-1, -1), 1, border_color),
        ('BACKGROUND', (0, -1), (-1, -1), header_bg),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))

    combined = Table([[e_table, d_table]], colWidths=[half_w + 4, half_w + 4], hAlign='CENTER')
    combined.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),
        ('RIGHTPADDING', (1, 0), (1, 0), 0),
        ('LEFTPADDING', (1, 0), (1, 0), 8),
    ]))
    elements.append(combined)
    elements.append(Spacer(1, 16))

    # ============ TOTAL NET PAYABLE (green box) ============
    net_title_style = ParagraphStyle('NTS', fontSize=9, fontName='Helvetica-Bold', textColor=dark_text)
    net_formula_style = ParagraphStyle('NFS', fontSize=7, textColor=light_text)
    net_big_style = ParagraphStyle('NBS', fontSize=18, fontName='Helvetica-Bold', textColor=dark_text, alignment=2)

    net_left = []
    net_left.append(Paragraph("TOTAL NET PAYABLE", net_title_style))
    net_left.append(Spacer(1, 2))
    net_left.append(Paragraph("Gross Earnings - Total Deductions", net_formula_style))

    net_box_data = [[net_left, Paragraph(f"Rs. {net_pay:,.2f}", net_big_style)]]
    net_box = Table(net_box_data, colWidths=[page_width * 0.55, page_width * 0.45])
    net_box.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), green_bg),
        ('BOX', (0, 0), (-1, -1), 1.5, green_border),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
    ]))
    elements.append(net_box)
    elements.append(Spacer(1, 10))

    # ============ AMOUNT IN WORDS ============
    try:
        net_int = int(net_pay)
        paise = int(round((net_pay - net_int) * 100))
        words = num2words(net_int, lang='en_IN').replace(',', '').title()
        if paise > 0:
            paise_words = num2words(paise, lang='en_IN').title()
            amount_words = f"Indian Rupee {words} and {paise_words} Paise Only"
        else:
            amount_words = f"Indian Rupee {words} Only"
    except Exception:
        amount_words = ""

    words_style = ParagraphStyle('WS', fontSize=8, textColor=medium_text, alignment=1)
    words_label = ParagraphStyle('WL', fontSize=7, textColor=light_text, alignment=1)
    elements.append(Paragraph("Amount In Words:", words_label))
    elements.append(Spacer(1, 2))
    elements.append(Paragraph(amount_words, words_style))
    elements.append(Spacer(1, 25))

    # ============ FOOTER ============
    elements.append(HRFlowable(width="100%", thickness=0.5, color=border_color))
    elements.append(Spacer(1, 6))
    footer_style = ParagraphStyle('FT', fontSize=7, alignment=1, textColor=light_text)
    elements.append(Paragraph("-- This document has been automatically generated by Sparkcurv HR Portal; therefore, a signature is not required. --", footer_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer

# Admin Payslip Routes
@admin_router.put("/employees/{employee_id}/salary")
async def set_employee_salary(employee_id: str, salary_data: SalaryUpdate, request: Request):
    await require_admin(request)
    result = await execute_query("UPDATE users SET basic_salary = %s WHERE id = %s", (salary_data.basic_salary, int(employee_id)))
    if result == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    return {"message": "Salary updated successfully"}

@admin_router.get("/employees/{employee_id}/salary")
async def get_employee_salary(employee_id: str, request: Request):
    await require_admin(request)
    emp = await execute_query("SELECT basic_salary FROM users WHERE id = %s", (int(employee_id),), fetch_one=True)
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    return {"basic_salary": emp.get("basic_salary", 0) or 0}

@admin_router.post("/payslip/generate")
async def generate_payslip(data: PayslipGenerate, request: Request):
    admin = await require_admin(request)
    employee = await execute_query("SELECT * FROM users WHERE id = %s", (int(data.employee_id),), fetch_one=True)
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    basic_salary = employee.get("basic_salary", 0) or 0
    if basic_salary <= 0:
        raise HTTPException(status_code=400, detail="Employee salary not set")

    existing = await execute_query("SELECT id FROM payslips WHERE employee_id = %s AND month = %s AND year = %s", (int(data.employee_id), data.month, data.year), fetch_one=True)
    if existing:
        raise HTTPException(status_code=400, detail="Payslip already exists for this month")

    deductions = []
    total_deductions = 0
    month_start = datetime(data.year, data.month, 1, tzinfo=timezone.utc)
    month_end = datetime(data.year + 1, 1, 1, tzinfo=timezone.utc) if data.month == 12 else datetime(data.year, data.month + 1, 1, tzinfo=timezone.utc)
    per_day_salary = basic_salary / WORKING_DAYS_PER_MONTH
    half_day_salary = per_day_salary / 2

    # Half-day deductions
    leave_deds = await execute_query(
        "SELECT amount FROM leave_deductions WHERE user_id = %s AND date >= %s AND date < %s",
        (int(data.employee_id), month_start.isoformat(), month_end.isoformat()), fetch_all=True
    )
    half_day_amount = sum(d.get("amount", 0) for d in (leave_deds or []))
    if half_day_amount > 0:
        amount = half_day_amount * half_day_salary
        deductions.append({"description": f"Half-day deduction ({half_day_amount} days)", "amount": round(amount, 2)})
        total_deductions += amount

    # LOP deductions
    lop_leaves = await execute_query(
        "SELECT days FROM leave_requests WHERE user_id = %s AND status = 'approved' AND leave_type = 'loss_of_pay' AND start_date >= %s AND start_date < %s",
        (int(data.employee_id), month_start.strftime("%Y-%m-%d"), month_end.strftime("%Y-%m-%d")), fetch_all=True
    )
    lop_days = sum(leave.get("days", 0) for leave in (lop_leaves or []))
    if lop_days > 0:
        lop_amount = lop_days * per_day_salary
        deductions.append({"description": f"Loss of Pay ({lop_days} days)", "amount": round(lop_amount, 2)})
        total_deductions += lop_amount

    # Custom deductions (PF, ESI, TDS, etc.)
    custom_deds = await execute_query(
        "SELECT * FROM custom_deductions WHERE user_id = %s AND is_active = 1",
        (int(data.employee_id),), fetch_all=True
    )
    for cd in (custom_deds or []):
        if cd.get("is_percentage") and cd.get("percentage", 0) > 0:
            ded_amount = round(basic_salary * cd["percentage"] / 100, 2)
        else:
            ded_amount = cd.get("amount", 0)
        if ded_amount > 0:
            deductions.append({"description": cd["deduction_name"], "amount": round(ded_amount, 2)})
            total_deductions += ded_amount

    month_names = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    net_pay = basic_salary - total_deductions

    payslip_id = await execute_query(
        """INSERT INTO payslips (employee_id, employee_name, employee_email, department, position, month, year, month_name, basic_salary, deduction_details, total_deductions, net_pay, generated_by, created_at)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
        (int(data.employee_id), employee["name"], employee.get("email", ""), employee.get("department", ""), employee.get("position", ""),
         data.month, data.year, month_names[data.month], basic_salary, json.dumps(deductions), round(total_deductions, 2), round(net_pay, 2),
         admin["name"], datetime.now(timezone.utc).isoformat()),
        last_id=True
    )

    return {"id": str(payslip_id), "employee_id": str(data.employee_id), "employee_name": employee["name"], "month": data.month, "year": data.year, "month_name": month_names[data.month], "basic_salary": basic_salary, "deduction_details": deductions, "total_deductions": round(total_deductions, 2), "net_pay": round(net_pay, 2), "generated_by": admin["name"], "created_at": datetime.now(timezone.utc).isoformat()}

@admin_router.get("/payslips")
async def get_all_payslips(request: Request, month: Optional[int] = None, year: Optional[int] = None):
    await require_admin(request)
    query = "SELECT * FROM payslips"
    conditions = []
    args = []
    if month:
        conditions.append("month = %s")
        args.append(month)
    if year:
        conditions.append("year = %s")
        args.append(year)
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY created_at DESC"
    rows = await execute_query(query, tuple(args) if args else None, fetch_all=True)
    result = []
    for ps in (rows or []):
        d = dict(ps)
        d["id"] = str(d.pop("id"))
        d["employee_id"] = str(d["employee_id"])
        if isinstance(d.get("deduction_details"), str):
            d["deduction_details"] = json.loads(d["deduction_details"])
        result.append(d)
    return result

@admin_router.delete("/payslips/{payslip_id}")
async def delete_payslip(payslip_id: str, request: Request):
    await require_admin(request)
    result = await execute_query("DELETE FROM payslips WHERE id = %s", (int(payslip_id),))
    if result == 0:
        raise HTTPException(status_code=404, detail="Payslip not found")
    return {"message": "Payslip deleted"}

# ============== SHIFT ROUTES ==============

@admin_router.get("/shifts")
async def get_shifts(request: Request):
    await require_admin(request)
    return {"default_start": DEFAULT_SHIFT_START, "default_end": DEFAULT_SHIFT_END}

@admin_router.put("/employees/{employee_id}/shift")
async def assign_shift(employee_id: str, shift_data: ShiftAssign, request: Request):
    await require_admin(request)
    import re
    time_re = re.compile(r"^\d{2}:\d{2}$")
    if not time_re.match(shift_data.start_time) or not time_re.match(shift_data.end_time):
        raise HTTPException(status_code=400, detail="Times must be in HH:MM format")
    employee = await execute_query("SELECT shift FROM users WHERE id = %s", (int(employee_id),), fetch_one=True)
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    if employee.get("shift") and employee.get("shift") != "":
        raise HTTPException(status_code=400, detail="Shift already assigned. Use change shift to update.")
    shift_str = f"{shift_data.start_time}-{shift_data.end_time}"
    await execute_query("UPDATE users SET shift = %s WHERE id = %s", (shift_str, int(employee_id)))
    return {"message": f"Shift {shift_data.start_time}–{shift_data.end_time} assigned", "shift": shift_str}

@admin_router.put("/employees/{employee_id}/shift/change")
async def change_shift(employee_id: str, shift_data: ShiftAssign, request: Request):
    await require_admin(request)
    import re
    time_re = re.compile(r"^\d{2}:\d{2}$")
    if not time_re.match(shift_data.start_time) or not time_re.match(shift_data.end_time):
        raise HTTPException(status_code=400, detail="Times must be in HH:MM format")
    employee = await execute_query("SELECT id FROM users WHERE id = %s", (int(employee_id),), fetch_one=True)
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    shift_str = f"{shift_data.start_time}-{shift_data.end_time}"
    await execute_query("UPDATE users SET shift = %s WHERE id = %s", (shift_str, int(employee_id)))
    return {"message": f"Shift updated to {shift_data.start_time}–{shift_data.end_time}", "shift": shift_str}
# ============== REPORTS ==============

@reports_router.get("/attendance/export")
async def export_attendance_report(request: Request, start_date: str, end_date: str, employee_id: Optional[str] = None):
    await require_admin(request)
    query = "SELECT * FROM attendance WHERE date >= %s AND date <= %s"
    args = [start_date, end_date]
    if employee_id:
        query += " AND user_id = %s"
        args.append(int(employee_id))
    query += " ORDER BY date ASC, user_name ASC"
    records = await execute_query(query, tuple(args), fetch_all=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="002FA7", end_color="002FA7", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ["Employee Name", "Date", "Clock In", "Clock Out", "Working Hours", "Break Time (min)", "Short Day", "Total Hours"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    total_hours_all = 0
    for row_num, record in enumerate(records or [], 2):
        clock_in = datetime.fromisoformat(record["clock_in"]).strftime("%I:%M %p") if record.get("clock_in") else "-"
        clock_out = datetime.fromisoformat(record["clock_out"]).strftime("%I:%M %p") if record.get("clock_out") else "-"
        wh = record.get("working_hours", 0) or 0
        total_hours_all += wh
        row_data = [record.get("user_name", ""), record.get("date", ""), clock_in, clock_out, f"{wh:.2f}" if wh else "-", record.get("total_break_minutes", 0), "Yes" if record.get("is_short_day") else "No", f"{wh:.2f}" if wh else "-"]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border

    for w, c in [('A', 20), ('B', 12), ('C', 12), ('D', 12), ('E', 15), ('F', 15), ('G', 12), ('H', 12)]:
        ws.column_dimensions[w].width = c

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=attendance_{start_date}_to_{end_date}.xlsx"})

# ============== POLICY ROUTES ==============

@policy_router.get("/list")
async def get_policies():
    """Get all company policies - accessible to all authenticated users"""
    rows = await execute_query("SELECT * FROM policies ORDER BY sort_order ASC, id ASC", fetch_all=True)
    result = []
    for r in (rows or []):
        d = dict(r)
        d["id"] = str(d.pop("id"))
        result.append(d)
    return result

@policy_router.post("/create")
async def create_policy(data: PolicyCreate, request: Request):
    await require_admin(request)
    pid = await execute_query(
        "INSERT INTO policies (title, category, content, icon, sort_order, created_at, updated_at) VALUES (%s, %s, %s, %s, %s, %s, %s)",
        (data.title, data.category, data.content, data.icon, data.sort_order, datetime.now(timezone.utc).isoformat(), datetime.now(timezone.utc).isoformat()),
        last_id=True
    )
    return {"id": str(pid), "title": data.title, "category": data.category, "content": data.content, "icon": data.icon, "sort_order": data.sort_order}

@policy_router.put("/{policy_id}")
async def update_policy(policy_id: str, data: PolicyUpdate, request: Request):
    await require_admin(request)
    update_dict = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_dict:
        raise HTTPException(status_code=400, detail="No update data")
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    set_clause = ", ".join([f"{k} = %s" for k in update_dict])
    values = list(update_dict.values()) + [int(policy_id)]
    result = await execute_query(f"UPDATE policies SET {set_clause} WHERE id = %s", tuple(values))
    if result == 0:
        raise HTTPException(status_code=404, detail="Policy not found")
    return {"message": "Policy updated"}

@policy_router.delete("/{policy_id}")
async def delete_policy(policy_id: str, request: Request):
    await require_admin(request)
    result = await execute_query("DELETE FROM policies WHERE id = %s", (int(policy_id),))
    if result == 0:
        raise HTTPException(status_code=404, detail="Policy not found")
    return {"message": "Policy deleted"}

# ============== HOLIDAY ROUTES ==============

@holidays_router.get("/list")
async def get_holidays():
    return HOLIDAYS

@holidays_router.get("/check/{date}")
async def check_date(date: str):
    weekend = is_weekend(date)
    holiday = is_holiday(date)
    return {"date": date, "is_weekend": weekend, "is_holiday": holiday, "holiday_name": get_holiday_name(date) if holiday else "", "is_working_day": not weekend and not holiday}

# ============== WFH ROUTES ==============

@wfh_router.post("/request")
async def create_wfh_request(wfh_data: WFHRequest, request: Request):
    user = await get_current_user(request)
    # Check if date is a working day
    if is_weekend(wfh_data.date) or is_holiday(wfh_data.date):
        raise HTTPException(status_code=400, detail="Cannot request WFH on weekends or holidays")
    # Check duplicate
    existing = await execute_query(
        "SELECT id FROM wfh_requests WHERE user_id = %s AND date = %s AND status != 'rejected'",
        (user["id"], wfh_data.date), fetch_one=True
    )
    if existing:
        raise HTTPException(status_code=400, detail="WFH request already exists for this date")
    # Check monthly limit
    month_start = wfh_data.date[:7] + "-01"
    month_num = int(wfh_data.date[5:7])
    year_num = int(wfh_data.date[:4])
    if month_num == 12:
        month_end = f"{year_num + 1}-01-01"
    else:
        month_end = f"{year_num}-{month_num + 1:02d}-01"
    used = await execute_query(
        "SELECT COUNT(*) as cnt FROM wfh_requests WHERE user_id = %s AND date >= %s AND date < %s AND status IN ('pending', 'approved')",
        (user["id"], month_start, month_end), fetch_one=True
    )
    wfh_limit = user.get("wfh_limit")  # None = no limit set by admin
    used_count = used["cnt"] if used else 0
    if wfh_limit is not None and used_count >= wfh_limit:
        raise HTTPException(status_code=400, detail=f"Monthly WFH limit reached ({wfh_limit} days). Contact admin to increase your limit.")

    wfh_id = await execute_query(
        """INSERT INTO wfh_requests (user_id, user_name, user_email, date, reason, status, created_at)
           VALUES (%s, %s, %s, %s, %s, 'pending', %s)""",
        (user["id"], user["name"], user.get("email", ""), wfh_data.date, wfh_data.reason, datetime.now(timezone.utc).isoformat()),
        last_id=True
    )
    return {"id": str(wfh_id), "user_id": str(user["id"]), "user_name": user["name"], "date": wfh_data.date, "reason": wfh_data.reason, "status": "pending", "created_at": datetime.now(timezone.utc).isoformat()}

@wfh_router.get("/my-requests")
async def get_my_wfh_requests(request: Request):
    user = await get_current_user(request)
    rows = await execute_query(
        "SELECT * FROM wfh_requests WHERE user_id = %s ORDER BY created_at DESC",
        (user["id"],), fetch_all=True
    )
    result = []
    for r in (rows or []):
        d = dict(r)
        d["id"] = str(d.pop("id"))
        d["user_id"] = str(d["user_id"])
        result.append(d)
    return result

@wfh_router.get("/balance")
async def get_wfh_balance(request: Request):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc)
    month_start = f"{now.year}-{now.month:02d}-01"
    if now.month == 12:
        month_end = f"{now.year + 1}-01-01"
    else:
        month_end = f"{now.year}-{now.month + 1:02d}-01"
    used = await execute_query(
        "SELECT COUNT(*) as cnt FROM wfh_requests WHERE user_id = %s AND date >= %s AND date < %s AND status IN ('pending', 'approved')",
        (user["id"], month_start, month_end), fetch_one=True
    )
    wfh_limit = user.get("wfh_limit")  # None = admin hasn't set a limit
    used_count = used["cnt"] if used else 0
    remaining = None if wfh_limit is None else max(0, wfh_limit - used_count)
    return {"limit": wfh_limit, "used": used_count, "remaining": remaining}

@wfh_router.delete("/{wfh_id}")
async def cancel_wfh_request(wfh_id: str, request: Request):
    user = await get_current_user(request)
    wfh_req = await execute_query("SELECT * FROM wfh_requests WHERE id = %s", (int(wfh_id),), fetch_one=True)
    if not wfh_req:
        raise HTTPException(status_code=404, detail="WFH request not found")
    if wfh_req["user_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    if wfh_req["status"] != "pending":
        raise HTTPException(status_code=400, detail="Can only cancel pending requests")
    await execute_query("DELETE FROM wfh_requests WHERE id = %s", (int(wfh_id),))
    return {"message": "WFH request cancelled"}

# Admin WFH management
@admin_router.get("/wfh-requests")
async def get_all_wfh_requests(request: Request, status: Optional[str] = None):
    user = await require_admin_or_manager(request)
    query = "SELECT * FROM wfh_requests"
    args = []
    conditions = []
    if status:
        conditions.append("status = %s")
        args.append(status)
    if user["role"] == "manager":
        conditions.append("user_id IN (SELECT id FROM users WHERE department = %s)")
        args.append(user.get("department", ""))
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY created_at DESC"
    rows = await execute_query(query, tuple(args) if args else None, fetch_all=True)
    result = []
    for r in (rows or []):
        d = dict(r)
        d["id"] = str(d.pop("id"))
        d["user_id"] = str(d["user_id"])
        result.append(d)
    return result

@admin_router.put("/wfh-requests/{wfh_id}")
async def review_wfh_request(wfh_id: str, request: Request, action: str):
    reviewer = await require_admin_or_manager(request)
    if action not in ("approve", "reject"):
        raise HTTPException(status_code=400, detail="Invalid action")
    wfh_req = await execute_query("SELECT * FROM wfh_requests WHERE id = %s", (int(wfh_id),), fetch_one=True)
    if not wfh_req:
        raise HTTPException(status_code=404, detail="WFH request not found")
    if wfh_req["status"] != "pending":
        raise HTTPException(status_code=400, detail="WFH request already processed")
    new_status = "approved" if action == "approve" else "rejected"
    await execute_query(
        "UPDATE wfh_requests SET status = %s, reviewed_by = %s, reviewed_at = %s WHERE id = %s",
        (new_status, reviewer["name"], datetime.now(timezone.utc).isoformat(), int(wfh_id))
    )
    return {"message": f"WFH request {new_status}"}

# ============== PAYROLL ROUTES ==============

# Custom Deductions CRUD
@admin_router.get("/deductions/{employee_id}")
async def get_employee_deductions(employee_id: str, request: Request):
    await require_admin_or_manager(request)
    rows = await execute_query(
        "SELECT * FROM custom_deductions WHERE user_id = %s ORDER BY deduction_name",
        (int(employee_id),), fetch_all=True
    )
    result = []
    for r in (rows or []):
        d = dict(r)
        d["id"] = str(d.pop("id"))
        d["user_id"] = str(d["user_id"])
        d["is_percentage"] = bool(d.get("is_percentage", 0))
        d["is_active"] = bool(d.get("is_active", 1))
        result.append(d)
    return result

@admin_router.post("/deductions")
async def add_custom_deduction(data: CustomDeductionCreate, request: Request):
    await require_admin(request)
    ded_id = await execute_query(
        """INSERT INTO custom_deductions (user_id, deduction_name, amount, is_percentage, percentage, is_active, created_at)
           VALUES (%s, %s, %s, %s, %s, %s, %s)""",
        (data.user_id, data.deduction_name, data.amount, 1 if data.is_percentage else 0,
         data.percentage, 1 if data.is_active else 0, datetime.now(timezone.utc).isoformat()),
        last_id=True
    )
    return {"id": str(ded_id), "user_id": str(data.user_id), "deduction_name": data.deduction_name,
            "amount": data.amount, "is_percentage": data.is_percentage, "percentage": data.percentage, "is_active": data.is_active}

@admin_router.delete("/deductions/{deduction_id}")
async def delete_custom_deduction(deduction_id: str, request: Request):
    await require_admin(request)
    await execute_query("DELETE FROM custom_deductions WHERE id = %s", (int(deduction_id),))
    return {"message": "Deduction deleted"}

@admin_router.put("/deductions/{deduction_id}/toggle")
async def toggle_deduction(deduction_id: str, request: Request):
    await require_admin(request)
    ded = await execute_query("SELECT is_active FROM custom_deductions WHERE id = %s", (int(deduction_id),), fetch_one=True)
    if not ded:
        raise HTTPException(status_code=404, detail="Deduction not found")
    new_status = 0 if ded["is_active"] else 1
    await execute_query("UPDATE custom_deductions SET is_active = %s WHERE id = %s", (new_status, int(deduction_id)))
    return {"is_active": bool(new_status)}

# Bulk Payroll Processing
@admin_router.post("/payroll/process")
async def process_bulk_payroll(data: BulkPayrollGenerate, request: Request):
    admin = await require_admin(request)
    employees = await execute_query(
        "SELECT * FROM users WHERE role != 'admin' AND basic_salary > 0", fetch_all=True
    )
    if not employees:
        raise HTTPException(status_code=400, detail="No employees with salary set")

    month_names = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    results = {"generated": 0, "skipped": 0, "errors": [], "total_gross": 0, "total_deductions": 0, "total_net": 0}

    for employee in employees:
        try:
            existing = await execute_query("SELECT id FROM payslips WHERE employee_id = %s AND month = %s AND year = %s",
                (employee["id"], data.month, data.year), fetch_one=True)
            if existing:
                results["skipped"] += 1
                continue

            basic_salary = employee.get("basic_salary", 0) or 0
            deductions = []
            total_deductions = 0
            month_start = datetime(data.year, data.month, 1, tzinfo=timezone.utc)
            month_end = datetime(data.year + 1, 1, 1, tzinfo=timezone.utc) if data.month == 12 else datetime(data.year, data.month + 1, 1, tzinfo=timezone.utc)
            per_day_salary = basic_salary / WORKING_DAYS_PER_MONTH
            half_day_salary = per_day_salary / 2

            # Half-day deductions
            leave_deds = await execute_query(
                "SELECT amount FROM leave_deductions WHERE user_id = %s AND date >= %s AND date < %s",
                (employee["id"], month_start.isoformat(), month_end.isoformat()), fetch_all=True
            )
            half_day_amount = sum(d.get("amount", 0) for d in (leave_deds or []))
            if half_day_amount > 0:
                amount = half_day_amount * half_day_salary
                deductions.append({"description": f"Half-day deduction ({half_day_amount} days)", "amount": round(amount, 2)})
                total_deductions += amount

            # LOP deductions
            lop_leaves = await execute_query(
                "SELECT days FROM leave_requests WHERE user_id = %s AND status = 'approved' AND leave_type = 'loss_of_pay' AND start_date >= %s AND start_date < %s",
                (employee["id"], month_start.strftime("%Y-%m-%d"), month_end.strftime("%Y-%m-%d")), fetch_all=True
            )
            lop_days = sum(leave.get("days", 0) for leave in (lop_leaves or []))
            if lop_days > 0:
                lop_amount = lop_days * per_day_salary
                deductions.append({"description": f"Loss of Pay ({lop_days} days)", "amount": round(lop_amount, 2)})
                total_deductions += lop_amount

            # Custom deductions (PF, ESI, TDS, etc.)
            custom_deds = await execute_query(
                "SELECT * FROM custom_deductions WHERE user_id = %s AND is_active = 1",
                (employee["id"],), fetch_all=True
            )
            for cd in (custom_deds or []):
                if cd.get("is_percentage") and cd.get("percentage", 0) > 0:
                    ded_amount = round(basic_salary * cd["percentage"] / 100, 2)
                else:
                    ded_amount = cd.get("amount", 0)
                if ded_amount > 0:
                    deductions.append({"description": cd["deduction_name"], "amount": round(ded_amount, 2)})
                    total_deductions += ded_amount

            net_pay = basic_salary - total_deductions

            await execute_query(
                """INSERT INTO payslips (employee_id, employee_name, employee_email, department, position, month, year, month_name, basic_salary, deduction_details, total_deductions, net_pay, generated_by, created_at)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (employee["id"], employee["name"], employee.get("email", ""), employee.get("department", ""),
                 employee.get("position", ""), data.month, data.year, month_names[data.month],
                 basic_salary, json.dumps(deductions), round(total_deductions, 2), round(net_pay, 2),
                 admin["name"], datetime.now(timezone.utc).isoformat()),
                last_id=True
            )
            results["generated"] += 1
            results["total_gross"] += basic_salary
            results["total_deductions"] += total_deductions
            results["total_net"] += net_pay
        except Exception as e:
            results["errors"].append(f"{employee['name']}: {str(e)}")

    # Save payroll run
    try:
        await execute_query(
            """INSERT INTO payroll_runs (month, year, total_employees, total_gross, total_deductions, total_net, processed_by, processed_at)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
               ON DUPLICATE KEY UPDATE total_employees=%s, total_gross=%s, total_deductions=%s, total_net=%s, processed_by=%s, processed_at=%s""",
            (data.month, data.year, results["generated"], round(results["total_gross"], 2), round(results["total_deductions"], 2), round(results["total_net"], 2), admin["name"], datetime.now(timezone.utc).isoformat(),
             results["generated"], round(results["total_gross"], 2), round(results["total_deductions"], 2), round(results["total_net"], 2), admin["name"], datetime.now(timezone.utc).isoformat())
        )
    except Exception:
        pass

    return results

# Also update single payslip generation to include custom deductions
# Payroll Analytics
@admin_router.get("/payroll/summary")
async def payroll_summary(request: Request, year: Optional[int] = None):
    await require_admin_or_manager(request)
    if not year:
        year = datetime.now(timezone.utc).year

    # Monthly summary
    monthly = await execute_query(
        "SELECT month, year, SUM(basic_salary) as total_gross, SUM(total_deductions) as total_ded, SUM(net_pay) as total_net, COUNT(*) as emp_count FROM payslips WHERE year = %s GROUP BY month, year ORDER BY month",
        (year,), fetch_all=True
    )

    # Department breakdown
    dept = await execute_query(
        "SELECT department, SUM(basic_salary) as total_gross, SUM(net_pay) as total_net, COUNT(*) as emp_count FROM payslips WHERE year = %s GROUP BY department ORDER BY total_gross DESC",
        (year,), fetch_all=True
    )

    # YTD totals
    ytd = await execute_query(
        "SELECT SUM(basic_salary) as total_gross, SUM(total_deductions) as total_ded, SUM(net_pay) as total_net, COUNT(*) as total_payslips FROM payslips WHERE year = %s",
        (year,), fetch_one=True
    )

    # Payroll runs
    runs = await execute_query(
        "SELECT * FROM payroll_runs WHERE year = %s ORDER BY month DESC",
        (year,), fetch_all=True
    )

    month_names = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    monthly_data = []
    for m in (monthly or []):
        monthly_data.append({
            "month": m["month"], "month_name": month_names[m["month"]], "year": m["year"],
            "total_gross": float(m["total_gross"] or 0), "total_deductions": float(m["total_ded"] or 0),
            "total_net": float(m["total_net"] or 0), "employee_count": m["emp_count"]
        })

    dept_data = []
    for d in (dept or []):
        dept_data.append({
            "department": d["department"] or "Unassigned",
            "total_gross": float(d["total_gross"] or 0), "total_net": float(d["total_net"] or 0),
            "employee_count": d["emp_count"]
        })

    runs_data = []
    for r in (runs or []):
        rd = dict(r)
        rd["id"] = str(rd.pop("id"))
        rd["month_name"] = month_names[rd["month"]]
        runs_data.append(rd)

    return {
        "year": year,
        "monthly": monthly_data,
        "departments": dept_data,
        "ytd": {
            "total_gross": float(ytd["total_gross"] or 0) if ytd else 0,
            "total_deductions": float(ytd["total_ded"] or 0) if ytd else 0,
            "total_net": float(ytd["total_net"] or 0) if ytd else 0,
            "total_payslips": ytd["total_payslips"] if ytd else 0
        },
        "runs": runs_data
    }

# Employee salary structure view
@payslip_router.get("/my-salary-structure")
async def get_my_salary_structure(request: Request):
    user = await get_current_user(request)
    basic_salary = user.get("basic_salary", 0) or 0
    basic = round(basic_salary * 0.50, 2)
    hra = round(basic_salary * 0.20, 2)
    medical = round(basic_salary * 0.045, 2)
    conveyance = round(basic_salary * 0.06, 2)
    special = round(basic_salary - basic - hra - medical - conveyance, 2)

    # Get custom deductions
    custom_deds = await execute_query(
        "SELECT * FROM custom_deductions WHERE user_id = %s AND is_active = 1",
        (user["id"],), fetch_all=True
    )
    deductions = []
    total_deductions = 0
    for cd in (custom_deds or []):
        if cd.get("is_percentage") and cd.get("percentage", 0) > 0:
            ded_amount = round(basic_salary * cd["percentage"] / 100, 2)
        else:
            ded_amount = cd.get("amount", 0)
        deductions.append({"name": cd["deduction_name"], "amount": round(ded_amount, 2), "is_percentage": bool(cd.get("is_percentage", 0)), "percentage": cd.get("percentage", 0)})
        total_deductions += ded_amount

    return {
        "gross_salary": basic_salary,
        "earnings": [
            {"name": "Basic", "amount": basic, "percentage": 50},
            {"name": "House Rent Allowance", "amount": hra, "percentage": 20},
            {"name": "Medical Allowance", "amount": medical, "percentage": 4.5},
            {"name": "Conveyance Allowance", "amount": conveyance, "percentage": 6},
            {"name": "Special Allowance", "amount": special, "percentage": 19.5},
        ],
        "deductions": deductions,
        "total_deductions": round(total_deductions, 2),
        "net_salary": round(basic_salary - total_deductions, 2)
    }

# ============== CHANGE REQUEST ROUTES ==============

CR_TYPES = ["Installation", "Maintenance", "Software", "Hardware", "Access", "Policy Change", "Salary Revision", "Leave Adjustment", "Shift Change", "General", "Other"]

@cr_router.post("/create")
async def create_cr(data: CRCreate, request: Request):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    metadata_json = json.dumps(data.metadata) if data.metadata else None
    cr_id = await execute_query(
        """INSERT INTO change_requests (requester_id, requester_name, title, description, cr_type, priority, status, manager_approval, admin_approval, metadata, created_at, updated_at)
           VALUES (%s, %s, %s, %s, %s, %s, 'pending', 'pending', 'pending', %s, %s, %s)""",
        (user["id"], user["name"], data.title, data.description, data.cr_type, data.priority, metadata_json, now, now),
        last_id=True
    )
    return {"id": str(cr_id), "title": data.title, "status": "pending", "message": "Change request created"}

@cr_router.get("/my-requests")
async def get_my_crs(request: Request):
    user = await get_current_user(request)
    crs = await execute_query(
        "SELECT * FROM change_requests WHERE requester_id = %s ORDER BY created_at DESC",
        (user["id"],), fetch_all=True
    )
    result = []
    for cr in (crs or []):
        d = dict(cr)
        d["id"] = str(d.pop("id"))
        d["requester_id"] = str(d["requester_id"])
        if d.get("manager_id"): d["manager_id"] = str(d["manager_id"])
        if d.get("admin_id"): d["admin_id"] = str(d["admin_id"])
        result.append(d)
    return result

@cr_router.delete("/{cr_id}")
async def delete_cr(cr_id: str, request: Request):
    user = await get_current_user(request)
    cr = await execute_query("SELECT * FROM change_requests WHERE id = %s", (int(cr_id),), fetch_one=True)
    if not cr:
        raise HTTPException(status_code=404, detail="CR not found")
    if cr["requester_id"] != user["id"] and user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    if cr["status"] not in ("pending",):
        raise HTTPException(status_code=400, detail="Can only delete pending CRs")
    await execute_query("DELETE FROM change_requests WHERE id = %s", (int(cr_id),))
    return {"message": "CR deleted"}

@cr_router.get("/types")
async def get_cr_types(request: Request):
    await get_current_user(request)
    return CR_TYPES

# Admin/Manager: list all CRs
@admin_router.get("/change-requests")
async def get_all_crs(request: Request):
    user = await require_admin_or_manager(request)
    if user["role"] == "admin":
        crs = await execute_query("SELECT * FROM change_requests ORDER BY created_at DESC", fetch_all=True)
    else:
        crs = await execute_query("SELECT * FROM change_requests WHERE status IN ('pending', 'manager_approved') ORDER BY created_at DESC", fetch_all=True)
    result = []
    for cr in (crs or []):
        d = dict(cr)
        d["id"] = str(d.pop("id"))
        d["requester_id"] = str(d["requester_id"])
        if d.get("manager_id"): d["manager_id"] = str(d["manager_id"])
        if d.get("admin_id"): d["admin_id"] = str(d["admin_id"])
        result.append(d)
    return result

# Manager approval (step 1)
@admin_router.put("/change-requests/{cr_id}/manager-action")
async def manager_action_cr(cr_id: str, request: Request, action: str = "approve", notes: str = ""):
    user = await require_admin_or_manager(request)
    cr = await execute_query("SELECT * FROM change_requests WHERE id = %s", (int(cr_id),), fetch_one=True)
    if not cr:
        raise HTTPException(status_code=404, detail="CR not found")
    if cr["manager_approval"] != "pending":
        raise HTTPException(status_code=400, detail="Manager already acted on this CR")

    now = datetime.now(timezone.utc).isoformat()
    if action == "approve":
        new_status = "manager_approved"
        mgr_approval = "approved"
    else:
        new_status = "rejected"
        mgr_approval = "rejected"

    await execute_query(
        "UPDATE change_requests SET status = %s, manager_approval = %s, manager_id = %s, manager_name = %s, manager_notes = %s, manager_action_at = %s, updated_at = %s WHERE id = %s",
        (new_status, mgr_approval, user["id"], user["name"], notes, now, now, int(cr_id))
    )
    return {"message": f"CR {action}d by manager", "status": new_status}

# Admin approval (step 2) with auto-apply
@admin_router.put("/change-requests/{cr_id}/admin-action")
async def admin_action_cr(cr_id: str, request: Request):
    admin = await require_admin(request)
    body = await request.json()
    action = body.get("action", "approve")
    notes = body.get("notes", "")
    apply_value = body.get("apply_value", "")

    cr = await execute_query("SELECT * FROM change_requests WHERE id = %s", (int(cr_id),), fetch_one=True)
    if not cr:
        raise HTTPException(status_code=404, detail="CR not found")
    if cr["admin_approval"] != "pending":
        raise HTTPException(status_code=400, detail="Admin already acted on this CR")

    now = datetime.now(timezone.utc).isoformat()
    if action == "approve":
        new_status = "approved"
        adm_approval = "approved"
    else:
        new_status = "rejected"
        adm_approval = "rejected"

    applied_changes = None
    # Auto-apply changes on approval
    if action == "approve" and apply_value:
        cr_type = cr.get("cr_type", "")
        requester_id = cr.get("requester_id")
        try:
            if cr_type == "Salary Revision":
                new_salary = float(apply_value)
                await execute_query("UPDATE users SET basic_salary = %s WHERE id = %s", (new_salary, requester_id))
                applied_changes = json.dumps({"type": "salary_revision", "new_value": new_salary})
            elif cr_type == "Leave Adjustment":
                new_leave = float(apply_value)
                await execute_query("UPDATE users SET casual_leave = %s WHERE id = %s", (new_leave, requester_id))
                applied_changes = json.dumps({"type": "leave_adjustment", "casual_leave": new_leave})
            elif cr_type == "Shift Change":
                new_shift = str(apply_value).strip()
                await execute_query("UPDATE users SET shift = %s WHERE id = %s", (new_shift, requester_id))
                applied_changes = json.dumps({"type": "shift_change", "new_shift": new_shift})
        except Exception as e:
            logger.warning(f"Auto-apply failed for CR {cr_id}: {e}")

    await execute_query(
        "UPDATE change_requests SET status = %s, admin_approval = %s, admin_id = %s, admin_name = %s, admin_notes = %s, admin_action_at = %s, applied_changes = %s, updated_at = %s WHERE id = %s",
        (new_status, adm_approval, admin["id"], admin["name"], notes, now, applied_changes, now, int(cr_id))
    )
    return {"message": f"CR {action}d by admin", "status": new_status, "applied": applied_changes is not None}

# ============== OFFICE SETTINGS ROUTES ==============

# Notification counts for manager/admin
@admin_router.get("/notifications")
async def get_notifications(request: Request):
    user = await require_admin_or_manager(request)
    pending_leaves = await execute_query("SELECT COUNT(*) as cnt FROM leave_requests WHERE status = 'pending'", fetch_one=True)
    pending_wfh = await execute_query("SELECT COUNT(*) as cnt FROM wfh_requests WHERE status = 'pending'", fetch_one=True)
    pending_permissions = await execute_query("SELECT COUNT(*) as cnt FROM permissions WHERE status = 'pending'", fetch_one=True)
    if user["role"] == "admin":
        pending_crs = await execute_query("SELECT COUNT(*) as cnt FROM change_requests WHERE status IN ('pending', 'manager_approved')", fetch_one=True)
    else:
        pending_crs = await execute_query("SELECT COUNT(*) as cnt FROM change_requests WHERE manager_approval = 'pending'", fetch_one=True)

    total = (pending_leaves["cnt"] or 0) + (pending_wfh["cnt"] or 0) + (pending_crs["cnt"] or 0) + (pending_permissions["cnt"] or 0)
    items = []
    if pending_leaves["cnt"]: items.append({"type": "leave", "count": pending_leaves["cnt"], "label": f"{pending_leaves['cnt']} pending leave request(s)"})
    if pending_wfh["cnt"]: items.append({"type": "wfh", "count": pending_wfh["cnt"], "label": f"{pending_wfh['cnt']} pending WFH request(s)"})
    if pending_crs["cnt"]: items.append({"type": "cr", "count": pending_crs["cnt"], "label": f"{pending_crs['cnt']} pending change request(s)"})
    if pending_permissions["cnt"]: items.append({"type": "permission", "count": pending_permissions["cnt"], "label": f"{pending_permissions['cnt']} pending permission request(s)"})
    return {"total": total, "items": items}

# Attendance heatmap — weekly on-time/late/absent
@admin_router.get("/attendance/heatmap")
async def get_attendance_heatmap(request: Request, weeks: int = 4):
    await require_admin_or_manager(request)
    today = datetime.now(timezone.utc)
    start = today - timedelta(days=weeks * 7)
    start_str = start.strftime("%Y-%m-%d")
    end_str = today.strftime("%Y-%m-%d")

    # Get all employees
    emps = await execute_query("SELECT id, name, employee_code, department FROM users WHERE role IN ('employee', 'manager') ORDER BY name", fetch_all=True)
    # Get attendance records in range
    records = await execute_query(
        "SELECT user_id, date, clock_in, working_hours, is_short_day FROM attendance WHERE date >= %s AND date <= %s",
        (start_str, end_str), fetch_all=True
    )

    # Build lookup: {user_id: {date: record}}
    att_map = {}
    for r in (records or []):
        uid = r["user_id"]
        if uid not in att_map:
            att_map[uid] = {}
        att_map[uid][r["date"]] = r

    # Generate date list (excluding weekends)
    dates = []
    d = start
    while d <= today:
        ds = d.strftime("%Y-%m-%d")
        if d.weekday() < 5:  # Mon-Fri
            dates.append(ds)
        d += timedelta(days=1)

    # Build heatmap data
    heatmap = []
    for emp in (emps or []):
        row = {"employee_id": str(emp["id"]), "name": emp["name"], "employee_code": emp.get("employee_code", ""), "department": emp.get("department", ""), "days": []}
        for dt in dates:
            rec = att_map.get(emp["id"], {}).get(dt)
            if rec:
                # Check if on-time: clock_in before 10:00 AM is on-time
                try:
                    cin = datetime.fromisoformat(rec["clock_in"])
                    clock_hour = cin.hour + cin.minute / 60
                    if clock_hour <= 10.0:
                        status = "ontime"
                    else:
                        status = "late"
                except Exception:
                    status = "present"
                if rec.get("is_short_day"):
                    status = "short"
                row["days"].append({"date": dt, "status": status, "hours": rec.get("working_hours")})
            else:
                row["days"].append({"date": dt, "status": "absent", "hours": None})
        heatmap.append(row)

    return {"dates": dates, "employees": heatmap}

# Break location alerts — breaks outside office geofence
@admin_router.get("/attendance/break-alerts")
async def get_break_alerts(request: Request, date: str = None):
    await require_admin_or_manager(request)
    if not date:
        date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    office = await get_office_settings()

    # Get today's attendance with breaks that have location
    att_records = await execute_query(
        "SELECT a.id, a.user_id, a.user_name, a.date FROM attendance a WHERE a.date = %s",
        (date,), fetch_all=True
    )
    alerts = []
    for att in (att_records or []):
        breaks_with_loc = await execute_query(
            "SELECT id, break_start, break_end, break_start_lat, break_start_lng FROM breaks WHERE attendance_id = %s AND break_start_lat IS NOT NULL",
            (att["id"],), fetch_all=True
        )
        for brk in (breaks_with_loc or []):
            dist = haversine_km(brk["break_start_lat"], brk["break_start_lng"], office["latitude"], office["longitude"])
            if dist > office["radius_km"]:
                alerts.append({
                    "attendance_id": str(att["id"]),
                    "user_id": str(att["user_id"]),
                    "user_name": att["user_name"],
                    "break_id": str(brk["id"]),
                    "break_start": brk["break_start"],
                    "break_end": brk["break_end"],
                    "latitude": brk["break_start_lat"],
                    "longitude": brk["break_start_lng"],
                    "distance_km": round(dist, 2),
                    "date": att["date"]
                })
    return alerts

@admin_router.get("/office-settings")
async def get_office_settings_api(request: Request):
    await require_admin(request)
    settings = await get_office_settings()
    return settings

@admin_router.put("/office-settings")
async def update_office_settings_api(data: OfficeSettingsUpdate, request: Request):
    await require_admin(request)
    now = datetime.now(timezone.utc).isoformat()
    existing = await execute_query("SELECT id FROM office_settings WHERE id = 1", fetch_one=True)
    if existing:
        await execute_query(
            "UPDATE office_settings SET latitude = %s, longitude = %s, radius_km = %s, office_name = %s, updated_at = %s WHERE id = 1",
            (data.latitude, data.longitude, data.radius_km, data.office_name, now)
        )
    else:
        await execute_query(
            "INSERT INTO office_settings (id, latitude, longitude, radius_km, office_name, updated_at) VALUES (1, %s, %s, %s, %s, %s)",
            (data.latitude, data.longitude, data.radius_km, data.office_name, now)
        )
    return {"message": "Office settings updated", "latitude": data.latitude, "longitude": data.longitude, "radius_km": data.radius_km}

# Employee location map data for admin
@admin_router.get("/attendance/locations")
async def get_attendance_locations(request: Request, date: str = None):
    await require_admin_or_manager(request)
    if not date:
        date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    records = await execute_query(
        """SELECT a.*, u.employee_code, u.department, u.position, u.avatar_url
           FROM attendance a JOIN users u ON a.user_id = u.id
           WHERE a.date = %s AND a.clock_in_lat IS NOT NULL
           ORDER BY a.clock_in DESC""",
        (date,), fetch_all=True
    )
    result = []
    for r in (records or []):
        d = dict(r)
        d["id"] = str(d.pop("id"))
        d["user_id"] = str(d["user_id"])
        result.append(d)
    return result

# ── Superadmin Reset ──────────────────────────────────────────────────────────
class ResetPortalRequest(BaseModel):
    password: str

@admin_router.post("/reset-portal")
async def reset_portal(body: ResetPortalRequest, request: Request):
    """Deletes ALL data except the superadmin user and Worker Tree. Requires admin password."""
    user = await require_admin(request)
    admin_id = user.get("id")

    # Verify password before proceeding
    db_user = await execute_query("SELECT password_hash FROM users WHERE id = %s", (int(admin_id),), fetch_one=True)
    if not db_user or not verify_password(body.password, db_user["password_hash"]):
        raise HTTPException(status_code=401, detail="Incorrect password. Reset cancelled.")

    tables_to_clear = [
        "attendance", "breaks", "leave_requests", "wfh_requests",
        "permissions", "payslips", "leave_deductions", "custom_deductions",
        "payroll_runs", "change_requests", "media",
    ]
    for table in tables_to_clear:
        try:
            await execute_query(f"DELETE FROM `{table}`", ())
        except Exception as e:
            logging.warning(f"Reset: could not clear {table}: {e}")

    try:
        await execute_query(
            "DELETE FROM users WHERE id != %s AND email NOT IN ('ponish.jino@sparkcurv.com', 'hr@sparkcurv.com')",
            (int(admin_id),)
        )
    except Exception as e:
        logging.warning(f"Reset: could not clear users: {e}")

    logging.info(f"Portal reset by admin id={admin_id}")
    return {"message": "Portal reset complete. All employee data cleared. Worker Tree and your account are preserved."}

# ============== INCLUDE ROUTERS ==============

api_router.include_router(auth_router)
api_router.include_router(attendance_router)
api_router.include_router(leave_router)
api_router.include_router(admin_router)
api_router.include_router(employees_router)
api_router.include_router(permission_router)
api_router.include_router(payslip_router)
api_router.include_router(reports_router)
api_router.include_router(holidays_router)
api_router.include_router(policy_router)
api_router.include_router(wfh_router)
api_router.include_router(cr_router)

@api_router.get("/")
async def root():
    return {"message": "HR Portal API"}

@api_router.get("/health")
async def health_check():
    """Health check endpoint to verify DB connection"""
    try:
        result = await execute_query("SELECT COUNT(*) as cnt FROM users", fetch_one=True)
        return {"status": "ok", "database": "connected", "users_count": result["cnt"] if result else 0}
    except Exception as e:
        return {"status": "error", "database": "disconnected", "error": str(e)}

app.include_router(api_router)

# CORS - Allow frontend URL
frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
allowed_origins = [frontend_url]
# Also allow without trailing slash
if frontend_url.endswith("/"):
    allowed_origins.append(frontend_url.rstrip("/"))
else:
    allowed_origins.append(frontend_url + "/")

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============== DATABASE INIT ==============

async def init_database():
    p = await get_pool()
    async with p.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    email VARCHAR(255) UNIQUE NOT NULL,
                    password_hash VARCHAR(255) NOT NULL,
                    name VARCHAR(255) NOT NULL,
                    role ENUM('admin', 'manager', 'employee') DEFAULT 'employee',
                    department VARCHAR(255) DEFAULT 'General',
                    position VARCHAR(255) DEFAULT 'Employee',
                    avatar_url TEXT,
                    created_at VARCHAR(64),
                    casual_leave FLOAT DEFAULT NULL,
                    sick_leave FLOAT DEFAULT NULL,
                    loss_of_pay FLOAT DEFAULT 0,
                    permission_hours FLOAT DEFAULT 2,
                    half_day_leave FLOAT DEFAULT 0,
                    shift VARCHAR(50) DEFAULT '',
                    basic_salary FLOAT DEFAULT 0,
                    employee_code VARCHAR(20) DEFAULT ''
                )
            """)
            # Add wfh_limit column if missing
            try:
                await cur.execute("SELECT wfh_limit FROM users LIMIT 1")
            except Exception:
                await cur.execute("ALTER TABLE users ADD COLUMN wfh_limit INT DEFAULT NULL")
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS wfh_requests (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    user_id INT NOT NULL,
                    user_name VARCHAR(255),
                    user_email VARCHAR(255),
                    date VARCHAR(20),
                    reason TEXT,
                    status VARCHAR(20) DEFAULT 'pending',
                    created_at VARCHAR(64),
                    reviewed_by VARCHAR(255),
                    reviewed_at VARCHAR(64),
                    INDEX idx_user_date (user_id, date)
                )
            """)
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS attendance (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    user_id INT NOT NULL,
                    user_name VARCHAR(255),
                    date VARCHAR(20),
                    clock_in VARCHAR(64),
                    clock_out VARCHAR(64),
                    total_break_minutes INT DEFAULT 0,
                    working_hours FLOAT,
                    is_short_day TINYINT DEFAULT 0,
                    clock_in_lat DOUBLE,
                    clock_in_lng DOUBLE,
                    clock_in_address TEXT,
                    clock_out_lat DOUBLE,
                    clock_out_lng DOUBLE,
                    clock_out_address TEXT,
                    location_type VARCHAR(20),
                    INDEX idx_user_date (user_id, date)
                )
            """)
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS breaks (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    attendance_id INT NOT NULL,
                    break_start VARCHAR(64),
                    break_end VARCHAR(64),
                    break_start_lat DOUBLE,
                    break_start_lng DOUBLE,
                    break_end_lat DOUBLE,
                    break_end_lng DOUBLE,
                    INDEX idx_attendance (attendance_id)
                )
            """)
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS leave_requests (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    user_id INT NOT NULL,
                    user_name VARCHAR(255),
                    user_email VARCHAR(255),
                    leave_type VARCHAR(50),
                    start_date VARCHAR(20),
                    end_date VARCHAR(20),
                    days FLOAT DEFAULT 0,
                    reason TEXT,
                    status VARCHAR(20) DEFAULT 'pending',
                    created_at VARCHAR(64),
                    reviewed_by VARCHAR(255),
                    reviewed_at VARCHAR(64),
                    is_half_day TINYINT DEFAULT 0,
                    INDEX idx_user_status (user_id, status)
                )
            """)
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS permissions (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    user_id INT NOT NULL,
                    user_name VARCHAR(255),
                    user_email VARCHAR(255),
                    duration_minutes INT,
                    reason TEXT,
                    date VARCHAR(20),
                    status VARCHAR(20) DEFAULT 'pending',
                    created_at VARCHAR(64),
                    reviewed_by VARCHAR(255),
                    reviewed_at VARCHAR(64),
                    INDEX idx_user_date (user_id, date)
                )
            """)
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS payslips (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    employee_id INT NOT NULL,
                    employee_name VARCHAR(255),
                    employee_email VARCHAR(255),
                    department VARCHAR(255),
                    position VARCHAR(255),
                    month INT,
                    year INT,
                    month_name VARCHAR(20),
                    basic_salary FLOAT,
                    deduction_details TEXT,
                    total_deductions FLOAT DEFAULT 0,
                    net_pay FLOAT DEFAULT 0,
                    generated_by VARCHAR(255),
                    created_at VARCHAR(64),
                    UNIQUE KEY uk_emp_month_year (employee_id, month, year)
                )
            """)
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS leave_deductions (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    user_id INT NOT NULL,
                    type VARCHAR(50),
                    amount FLOAT,
                    reason TEXT,
                    date VARCHAR(64)
                )
            """)
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS custom_deductions (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    user_id INT NOT NULL,
                    deduction_name VARCHAR(100) NOT NULL,
                    amount FLOAT DEFAULT 0,
                    is_percentage TINYINT DEFAULT 0,
                    percentage FLOAT DEFAULT 0,
                    is_active TINYINT DEFAULT 1,
                    created_at VARCHAR(64),
                    INDEX idx_user (user_id)
                )
            """)
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS payroll_runs (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    month INT NOT NULL,
                    year INT NOT NULL,
                    total_employees INT DEFAULT 0,
                    total_gross FLOAT DEFAULT 0,
                    total_deductions FLOAT DEFAULT 0,
                    total_net FLOAT DEFAULT 0,
                    processed_by VARCHAR(255),
                    processed_at VARCHAR(64),
                    UNIQUE KEY uk_month_year (month, year)
                )
            """)
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS policies (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    title VARCHAR(255) NOT NULL,
                    category VARCHAR(100),
                    content TEXT,
                    icon VARCHAR(50) DEFAULT 'article',
                    sort_order INT DEFAULT 0,
                    created_at VARCHAR(64),
                    updated_at VARCHAR(64)
                )
            """)
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS change_requests (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    requester_id INT NOT NULL,
                    requester_name VARCHAR(255),
                    title VARCHAR(255) NOT NULL,
                    description TEXT,
                    cr_type VARCHAR(100) DEFAULT 'General',
                    priority VARCHAR(20) DEFAULT 'medium',
                    status VARCHAR(30) DEFAULT 'pending',
                    manager_approval VARCHAR(20) DEFAULT 'pending',
                    manager_id INT,
                    manager_name VARCHAR(255),
                    manager_notes TEXT,
                    manager_action_at VARCHAR(64),
                    admin_approval VARCHAR(20) DEFAULT 'pending',
                    admin_id INT,
                    admin_name VARCHAR(255),
                    admin_notes TEXT,
                    admin_action_at VARCHAR(64),
                    metadata TEXT,
                    applied_changes TEXT,
                    created_at VARCHAR(64),
                    updated_at VARCHAR(64),
                    INDEX idx_requester (requester_id),
                    INDEX idx_status (status)
                )
            """)
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS office_settings (
                    id INT PRIMARY KEY DEFAULT 1,
                    latitude DOUBLE NOT NULL,
                    longitude DOUBLE NOT NULL,
                    radius_km FLOAT DEFAULT 0.5,
                    office_name VARCHAR(255) DEFAULT 'Office',
                    updated_at VARCHAR(64)
                )
            """)
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS org_chart (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    parent_id INT,
                    employee_name VARCHAR(255) NOT NULL,
                    job_title VARCHAR(255),
                    image_url TEXT,
                    description TEXT,
                    level_num INT DEFAULT 0,
                    sort_order INT DEFAULT 0,
                    created_at VARCHAR(64)
                )
            """)
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS org_levels (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    level_num INT NOT NULL,
                    label VARCHAR(100) NOT NULL DEFAULT 'Level',
                    updated_at VARCHAR(64),
                    UNIQUE KEY uk_level_num (level_num)
                )
            """)
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS role_permissions (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    role VARCHAR(20) NOT NULL,
                    feature_key VARCHAR(100) NOT NULL,
                    enabled TINYINT DEFAULT 1,
                    UNIQUE KEY uk_role_feature (role, feature_key)
                )
            """)
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS media (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    media_id VARCHAR(64) NOT NULL,
                    filename VARCHAR(255) NOT NULL,
                    content_type VARCHAR(64) NOT NULL DEFAULT 'image/jpeg',
                    data MEDIUMBLOB NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE KEY uk_media_id (media_id)
                )
            """)

@app.on_event("startup")
async def startup():
    import traceback
    try:
        logger.info("Connecting to MySQL...")
        logger.info(f"  Host: {os.environ.get('MYSQL_HOST', 'localhost')}")
        logger.info(f"  Port: {os.environ.get('MYSQL_PORT', '3306')}")
        logger.info(f"  Database: {os.environ.get('MYSQL_DB', 'hr_portal')}")
        logger.info(f"  User: {os.environ.get('MYSQL_USER', 'hruser')}")

        await init_database()
        logger.info("All tables created/verified successfully")

        # Migration: Add employee_code column if missing
        try:
            await execute_query("SELECT employee_code FROM users LIMIT 1", fetch_one=True)
        except Exception:
            try:
                await execute_query("ALTER TABLE users ADD COLUMN employee_code VARCHAR(20) DEFAULT ''")
                logger.info("Added employee_code column to users table")
            except Exception as e:
                logger.warning(f"Could not add employee_code column: {e}")

        # Migration: Add is_half_day to leave_requests and change days to FLOAT
        try:
            await execute_query("SELECT is_half_day FROM leave_requests LIMIT 1", fetch_one=True)
        except Exception:
            try:
                await execute_query("ALTER TABLE leave_requests ADD COLUMN is_half_day TINYINT DEFAULT 0")
                logger.info("Added is_half_day column to leave_requests")
            except Exception as e:
                logger.warning(f"Could not add is_half_day column: {e}")
        try:
            await execute_query("ALTER TABLE leave_requests MODIFY COLUMN days FLOAT DEFAULT 0")
        except Exception:
            pass

        # Migration: Add wfh_limit to users
        try:
            await execute_query("SELECT wfh_limit FROM users LIMIT 1", fetch_one=True)
        except Exception:
            try:
                await execute_query("ALTER TABLE users ADD COLUMN wfh_limit INT DEFAULT NULL")
                logger.info("Added wfh_limit column to users")
            except Exception as e:
                logger.warning(f"Could not add wfh_limit column: {e}")

        # Migration: Add location columns to attendance
        for col in ["clock_in_lat DOUBLE", "clock_in_lng DOUBLE", "clock_in_address TEXT", "clock_out_lat DOUBLE", "clock_out_lng DOUBLE", "clock_out_address TEXT", "location_type VARCHAR(20)"]:
            col_name = col.split()[0]
            try:
                await execute_query(f"SELECT {col_name} FROM attendance LIMIT 1", fetch_one=True)
            except Exception:
                try:
                    await execute_query(f"ALTER TABLE attendance ADD COLUMN {col}")
                    logger.info(f"Added {col_name} to attendance")
                except Exception:
                    pass

        # Migration: Add location columns to breaks
        for col in ["break_start_lat DOUBLE", "break_start_lng DOUBLE", "break_end_lat DOUBLE", "break_end_lng DOUBLE"]:
            col_name = col.split()[0]
            try:
                await execute_query(f"SELECT {col_name} FROM breaks LIMIT 1", fetch_one=True)
            except Exception:
                try:
                    await execute_query(f"ALTER TABLE breaks ADD COLUMN {col}")
                    logger.info(f"Added {col_name} to breaks")
                except Exception:
                    pass

        # Migration: Add metadata and applied_changes to change_requests
        for col in ["metadata TEXT", "applied_changes TEXT"]:
            col_name = col.split()[0]
            try:
                await execute_query(f"SELECT {col_name} FROM change_requests LIMIT 1", fetch_one=True)
            except Exception:
                try:
                    await execute_query(f"ALTER TABLE change_requests ADD COLUMN {col}")
                    logger.info(f"Added {col_name} to change_requests")
                except Exception as e:
                    logger.warning(f"Could not add {col_name} to change_requests: {e}")

        # Migration: Add level_num to org_chart
        try:
            await execute_query("SELECT level_num FROM org_chart LIMIT 1", fetch_one=True)
        except Exception:
            try:
                await execute_query("ALTER TABLE org_chart ADD COLUMN level_num INT DEFAULT 0")
                logger.info("Added level_num to org_chart")
            except Exception as e:
                logger.warning(f"Could not add level_num to org_chart: {e}")

        # Backfill employee_code for existing users without one
        try:
            users_without_code = await execute_query("SELECT id FROM users WHERE employee_code IS NULL OR employee_code = '' ORDER BY id", fetch_all=True)
            if users_without_code:
                max_code = await execute_query("SELECT employee_code FROM users WHERE employee_code LIKE 'SC%' ORDER BY employee_code DESC LIMIT 1", fetch_one=True)
                next_num = 24001
                if max_code and max_code["employee_code"]:
                    try:
                        next_num = int(max_code["employee_code"][2:]) + 1
                    except ValueError:
                        pass
                for u in users_without_code:
                    code = f"SC{next_num}"
                    await execute_query("UPDATE users SET employee_code = %s WHERE id = %s", (code, u["id"]))
                    next_num += 1
                logger.info(f"Backfilled employee_code for {len(users_without_code)} users")
        except Exception as e:
            logger.warning(f"Employee code backfill skipped: {e}")

        # Seed admin
        admin_email = os.environ.get("ADMIN_EMAIL", "admin@hrportal.com")
        admin_password = os.environ.get("ADMIN_PASSWORD", "Admin@123")

        existing = await execute_query("SELECT id, password_hash FROM users WHERE email = %s", (admin_email,), fetch_one=True)
        if existing is None:
            hashed = hash_password(admin_password)
            try:
                await execute_query(
                    """INSERT INTO users (email, password_hash, name, role, department, position, avatar_url, created_at, casual_leave, sick_leave, loss_of_pay, employee_code)
                       VALUES (%s, %s, 'Admin', 'admin', 'Administration', 'System Admin', '', %s, 12, 3, 0, 'SC24001')""",
                    (admin_email, hashed, datetime.now(timezone.utc).isoformat())
                )
            except Exception:
                await execute_query(
                    """INSERT INTO users (email, password_hash, name, role, department, position, avatar_url, created_at, casual_leave, sick_leave, loss_of_pay)
                       VALUES (%s, %s, 'Admin', 'admin', 'Administration', 'System Admin', '', %s, 12, 3, 0)""",
                    (admin_email, hashed, datetime.now(timezone.utc).isoformat())
                )
            logger.info(f"Admin user created: {admin_email} / {admin_password}")
        elif not verify_password(admin_password, existing["password_hash"]):
            await execute_query("UPDATE users SET password_hash = %s WHERE email = %s", (hash_password(admin_password), admin_email))
            logger.info("Admin password updated")
        else:
            logger.info(f"Admin user exists: {admin_email}")

        # Verify admin can login
        admin_check = await execute_query("SELECT id, email, role FROM users WHERE email = %s AND role = 'admin'", (admin_email,), fetch_one=True)
        if admin_check:
            logger.info(f"Admin login ready - Email: {admin_email}, Password: {admin_password}")
        else:
            logger.error("CRITICAL: Admin user not found after seeding!")

        # Seed second superadmin: hr@sparkcurv.com
        hr_admins = [
            {"email": "hr@sparkcurv.com", "password": "hr@2024", "name": "HR Admin", "code": "SC24002"},
        ]
        for ha in hr_admins:
            ex = await execute_query("SELECT id, password_hash FROM users WHERE email = %s", (ha["email"],), fetch_one=True)
            if ex is None:
                hashed2 = hash_password(ha["password"])
                try:
                    await execute_query(
                        """INSERT INTO users (email, password_hash, name, role, department, position, avatar_url, created_at, casual_leave, sick_leave, loss_of_pay, employee_code)
                           VALUES (%s, %s, %s, 'admin', 'Administration', 'HR Admin', '', %s, 12, 3, 0, %s)""",
                        (ha["email"], hashed2, ha["name"], datetime.now(timezone.utc).isoformat(), ha["code"])
                    )
                    logger.info(f"HR Admin created: {ha['email']}")
                except Exception as e2:
                    logger.warning(f"HR Admin seed fallback: {e2}")
                    await execute_query(
                        """INSERT INTO users (email, password_hash, name, role, department, position, avatar_url, created_at, casual_leave, sick_leave, loss_of_pay)
                           VALUES (%s, %s, %s, 'admin', 'Administration', 'HR Admin', '', %s, 12, 3, 0)""",
                        (ha["email"], hashed2, ha["name"], datetime.now(timezone.utc).isoformat())
                    )
            elif not verify_password(ha["password"], ex["password_hash"]):
                await execute_query("UPDATE users SET password_hash = %s WHERE email = %s", (hash_password(ha["password"]), ha["email"]))
                logger.info(f"HR Admin password synced: {ha['email']}")
            else:
                logger.info(f"HR Admin exists: {ha['email']}")

        # Seed default policies if none exist
        policy_count = await execute_query("SELECT COUNT(*) as cnt FROM policies", fetch_one=True)
        if policy_count and policy_count["cnt"] == 0:
            now = datetime.now(timezone.utc).isoformat()
            default_policies = [
                ("Leave Policy", "Leave", "Casual Leave: 12 days per year\nSick Leave: 3 days per year\nLoss of Pay (LOP): Available when paid leaves are exhausted. Salary will be deducted proportionally for each LOP day taken.\n\nLeave requests must be submitted in advance and approved by your Manager or Admin. Weekends (Saturday & Sunday) and public holidays are excluded from leave day calculations.", "calendar", 1),
                ("Working Hours", "Attendance", "Standard working hours: 8 hours 30 minutes per day\nMinimum required: 8 hours per day\n\nEmployees working less than 8 hours will be marked as a 'Short Day'. Every 3 short days will result in an automatic deduction of 0.5 day casual leave.", "clock", 2),
                ("Break Policy", "Attendance", "Maximum break time: 30 minutes per day\nBreaks can be taken in multiple sessions within the allowed limit.\n\nBreak time is deducted from total working hours. Exceeding the 30-minute break limit is not permitted.", "coffee", 3),
                ("Shift Timings", "Shift", "General Shift: 09:30 AM - 05:30 PM\nMorning Shift: 04:00 AM - 12:00 PM\nAfternoon Shift: 12:00 PM - 08:00 PM\nNight Shift: 08:00 PM - 04:00 AM\n\nShift assignments are managed by the Admin. Once assigned, shift changes require Admin approval.", "moon", 4),
                ("Permission Hours", "Permission", "Monthly allowance: 2 hours\nMaximum per use: 1 hour\n\nPermission requests must be submitted with a valid reason. Unused permission hours do not carry forward to the next month.", "hourglass", 5),
                ("Holiday Policy", "Holiday", "The company observes 12 public holidays per year.\nSaturday and Sunday are weekly holidays for all employees.\n\nEmployees cannot clock in on weekends or public holidays. The complete holiday list is available in the Holidays tab.", "flag", 6),
                ("Payslip", "Payroll", "Payslips are generated monthly by the Admin.\nDeductions include: Half-day deductions for short working days, Loss of Pay deductions.\n\nEmployees can download their payslips in PDF format from the Payslips section.", "receipt", 7),
            ]
            for title, category, content, icon, order in default_policies:
                await execute_query(
                    "INSERT INTO policies (title, category, content, icon, sort_order, created_at, updated_at) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                    (title, category, content, icon, order, now, now)
                )
            logger.info("Default company policies seeded")

    except Exception as e:
        logger.error(f"STARTUP ERROR: {str(e)}")
        logger.error(traceback.format_exc())
        logger.error("Make sure MySQL is running and .env has correct MYSQL_HOST, MYSQL_PORT, MYSQL_USER, MYSQL_PASSWORD, MYSQL_DB")

@app.on_event("shutdown")
async def shutdown():
    global pool
    if pool:
        pool.close()
        await pool.wait_closed()